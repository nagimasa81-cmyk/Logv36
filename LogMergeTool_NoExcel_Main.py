# -*- coding: utf-8 -*-
"""
Log Merge Tool - No Excel / No COM final version
Target: Windows 11 / Python 3.13 / latest PySide6 / latest Nuitka

This application does not start Microsoft Excel, does not run VBA, and does not use COM.
It parses log files in Python and writes an .xlsx report using openpyxl.
"""
from __future__ import annotations

import csv
import json
import fnmatch
import os
import re
import sys
import traceback
import tempfile
import zipfile
import xml.etree.ElementTree as ET
import concurrent.futures
import multiprocessing
import threading
import time
import sqlite3
from dataclasses import dataclass, asdict
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional, Iterable, Any

def startup_log_path() -> Path:
    base = os.environ.get("LOCALAPPDATA") or tempfile.gettempdir()
    p = Path(base) / "LogMergeTool_NoExcel"
    p.mkdir(parents=True, exist_ok=True)
    return p / "startup_error.log"


def write_startup_log(text: str) -> None:
    try:
        startup_log_path().write_text(text, encoding="utf-8")
    except Exception:
        pass


try:
    from PySide6.QtCore import Qt, QDate, QThread, Signal, QSettings, QAbstractTableModel, QModelIndex
    from PySide6.QtWidgets import (
        QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
        QFileDialog, QDateEdit, QCheckBox, QTextEdit, QComboBox, QMessageBox,
        QProgressBar, QGroupBox, QFormLayout, QLineEdit, QTableView, QSplitter,
        QSpinBox, QHeaderView, QDialog, QTableWidget, QTableWidgetItem, QDialogButtonBox, QInputDialog, QProgressDialog
    )
except Exception:
    write_startup_log("PySide6 import failed.\n\n" + traceback.format_exc())
    raise

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.formatting.rule import FormulaRule
except Exception:
    write_startup_log("openpyxl import failed.\n\n" + traceback.format_exc())
    raise

APP_TITLE = "Log Merge Tool - No Excel"
APP_VERSION = "2.36.0-noexcel-v36-plugin-manager"

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

# File classification is intentionally permissive so old field logs are not missed.
FILE_TYPE_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    ("WS", re.compile(r"^(?:20\d{2}_[A-Za-z]{3}_\d{2}_\d{2}_\d{2}_\d{2}(?:\(\d+\))?\.Log|WaterSystem.*\.(?:txt|log))$", re.I)),
    ("LAIS", re.compile(r"^lais.*\.(?:log|txt)$", re.I)),
    ("GESYS", re.compile(r"^gesys.*\.(?:log|txt)$", re.I)),
    ("MRSERVER", re.compile(r"^(?:mrserver|mr_server|mr-server).*\.(?:log|txt)$", re.I)),
    ("CSA", re.compile(r"^csa.*\.(?:log|txt)$", re.I)),
    ("PSC", re.compile(r"^psc(?:[._-].*)?\.(?:log|txt)$", re.I)),
    ("CGA", re.compile(r"^cga.*\.(?:log|txt)$", re.I)),
]


def plugins_dir() -> Path:
    p = app_dir() / "plugins"
    p.mkdir(parents=True, exist_ok=True)
    return p


def plugin_manifest_paths() -> list[Path]:
    root = plugins_dir()
    paths = []
    for p in root.iterdir():
        if p.is_dir() and (p / "manifest.json").exists():
            paths.append(p / "manifest.json")
    return sorted(paths, key=lambda x: x.parent.name.lower())


def load_file_type_plugins() -> list[dict[str, Any]]:
    plugins: list[dict[str, Any]] = []
    for manifest in plugin_manifest_paths():
        try:
            data = json.loads(manifest.read_text(encoding="utf-8"))
            if not data.get("id") or not data.get("display_name"):
                continue
            data["_path"] = str(manifest.parent)
            patterns = data.get("patterns") or data.get("file_patterns") or []
            if isinstance(patterns, str):
                patterns = [patterns]
            data["patterns"] = [str(x) for x in patterns if str(x).strip()]
            data["mode"] = data.get("mode") or ["merge", "import"]
            data["enabled"] = bool(data.get("enabled", True))
            if data["enabled"]:
                plugins.append(data)
        except Exception:
            continue
    return plugins


def plugin_type_patterns() -> list[tuple[str, re.Pattern[str]]]:
    out: list[tuple[str, re.Pattern[str]]] = []
    for plg in load_file_type_plugins():
        pid = str(plg.get("id", "")).strip()
        for pat in plg.get("patterns", []):
            try:
                out.append((pid, re.compile(fnmatch.translate(pat), re.I)))
            except Exception:
                pass
    return out


def validate_plugin_zip(zip_path: Path) -> tuple[bool, str, Optional[dict[str, Any]]]:
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            names = set(zf.namelist())
            if "manifest.json" not in names:
                return False, "manifest.json is missing.", None
            manifest = json.loads(zf.read("manifest.json").decode("utf-8-sig"))
            if not manifest.get("id"):
                return False, "manifest.json requires id.", None
            if not manifest.get("display_name"):
                return False, "manifest.json requires display_name.", None
            patterns = manifest.get("patterns") or manifest.get("file_patterns")
            if not patterns:
                return False, "manifest.json requires patterns or file_patterns.", None
            return True, "OK", manifest
    except Exception as exc:
        return False, str(exc), None


def install_plugin_zip(zip_path: Path) -> tuple[bool, str]:
    ok, msg, manifest = validate_plugin_zip(zip_path)
    if not ok or not manifest:
        return False, msg
    pid = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(manifest["id"]).strip())
    dest = plugins_dir() / pid
    if dest.exists():
        shutil.rmtree(dest, ignore_errors=True)
    dest.mkdir(parents=True, exist_ok=True)
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(dest)
        return True, f"Installed plugin: {manifest.get('display_name')} ({manifest.get('version','')})"
    except Exception as exc:
        shutil.rmtree(dest, ignore_errors=True)
        return False, str(exc)


def is_review_file(name: str) -> bool:
    """Review Import target: review.out, review.out.ar, or review.out.* only."""
    n = name.strip().lower()
    return n == "review.out" or n.startswith("review.out.")

WS_FILENAME_RE = re.compile(
    r"^(?P<y>20\d{2})_(?P<mon>[A-Za-z]{3})_(?P<d>\d{2})_(?P<h>\d{2})_(?P<mi>\d{2})_(?P<s>\d{2})(?:\(\d+\))?\.Log$",
    re.I,
)

FULL_DATETIME_PATTERNS: list[re.Pattern[str]] = [
    # 2026-07-04 14:33:04.149 / 2026/07/04 14:33:04:149
    re.compile(r"(?P<y>20\d{2})[-/](?P<m>\d{1,2})[-/](?P<d>\d{1,2})[ T_]+(?P<h>\d{1,2}):(?P<mi>\d{1,2})(?::(?P<s>\d{1,2})(?:(?:\.|:)(?P<ms>\d{1,6}))?)?"),
    # 07/04/2026 14:33:04.149
    re.compile(r"(?P<m>\d{1,2})/(?P<d>\d{1,2})/(?P<y>20\d{2})[ T_]+(?P<h>\d{1,2}):(?P<mi>\d{1,2})(?::(?P<s>\d{1,2})(?:(?:\.|:)(?P<ms>\d{1,6}))?)?"),
    # 04-Jul-2026 14:33:04.149
    re.compile(r"(?P<d>\d{1,2})[- ](?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ](?P<y>20\d{2})[ T_]+(?P<h>\d{1,2}):(?P<mi>\d{1,2})(?::(?P<s>\d{1,2})(?:(?:\.|:)(?P<ms>\d{1,6}))?)?", re.I),
    # Jul 04 2026 14:33:04.149
    re.compile(r"(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/ ](?P<d>\d{1,2})[-/ ,]+(?P<y>20\d{2})[ T_]+(?P<h>\d{1,2}):(?P<mi>\d{1,2})(?::(?P<s>\d{1,2})(?:(?:\.|:)(?P<ms>\d{1,6}))?)?", re.I),
]

DATE_ONLY_PATTERNS: list[re.Pattern[str]] = [
    re.compile(r"(?P<y>20\d{2})[-/](?P<m>\d{1,2})[-/](?P<d>\d{1,2})"),
    re.compile(r"(?P<m>\d{1,2})/(?P<d>\d{1,2})/(?P<y>20\d{2})"),
    re.compile(r"(?P<d>\d{1,2})[- ](?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ](?P<y>20\d{2})", re.I),
    re.compile(r"(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/ ](?P<d>\d{1,2})[-/ ,]+(?P<y>20\d{2})", re.I),
]

# WS and many system logs put the record time at the very beginning of each line.
# Example: 14:33:04:149 Ent Dbg 4992 ...
TIME_AT_LINE_START_RE = re.compile(r"^\s*(?P<h>\d{1,2}):(?P<mi>\d{2}):(?P<s>\d{2})(?:(?:\.|:)(?P<ms>\d{1,6}))?\b")

LEVEL_RE = re.compile(r"\b(ERROR|ERR|WARN(?:ING)?|INFO|DEBUG|TRACE|FATAL|ALARM)\b", re.I)

DEFAULT_CSA_RULES = [
    {"keyword": "Watch Dog Error", "message": "Watch Dog Error"},
    {"keyword": "Vacuum Failure", "message": "Vacuum Failure"},
]

DEFAULT_SITE_MAP = [
    {"serial": "4278", "site": "Yokosuka"},
    {"serial": "4259", "site": "CIM"},
    {"serial": "4257", "site": "Tsukazaki"},
    {"serial": "4087", "site": "Kyokuto"},
    {"serial": "4048", "site": "Kitano"},
    {"serial": "4083", "site": "Kashiwaba"},
    {"serial": "4078", "site": "Fukuoka"},
    {"serial": "4065", "site": "Miyagi"},
    {"serial": "4073", "site": "Atsuchi"},
]

@dataclass
class LogRecord:
    timestamp: Optional[datetime]
    source_type: str
    filename: str
    line_no: int
    level: str
    category: str
    message: str
    raw: str

@dataclass
class RunOptions:
    source_folder: str
    output_folder: str
    recursive: bool
    use_start: bool
    use_end: bool
    start_date: Optional[str]
    end_date: Optional[str]
    serial: str
    site: str
    include_ws: bool
    include_cga: bool
    include_csa: bool
    include_mrserver: bool
    include_gesys: bool
    include_lais: bool
    include_psc: bool
    include_unknown: bool
    turbo_csv: bool = True
    summary_xlsx_only: bool = True
    fast_xlsx_mode: bool = True
    csv_summary_mode: bool = True
    skip_merged1: bool = True
    timestamp_as_text: bool = True
    disable_styles_large: bool = True
    worker_count: int = 0
    noise_enabled_merge: bool = False
    noise_exclude_output: bool = False
    noise_learning: bool = True
    selected_files: Optional[list[str]] = None
    plugin_types: Optional[list[str]] = None

@dataclass
class RunResult:
    scanned_files: int = 0
    parsed_files: int = 0
    skipped_files: int = 0
    total_records: int = 0
    included_records: int = 0
    csa_error_count: int = 0
    output_path: str = ""


def app_dir() -> Path:
    """Return the folder that contains runtime data files.

    In Nuitka onefile mode, bundled data files are extracted beside __file__,
    while sys.argv[0] points to the outer EXE path.  Using __file__ keeps
    csa_error_rules.json and site_serial_map.json readable in both source and
    built EXE modes.
    """
    try:
        return Path(__file__).resolve().parent
    except Exception:
        return Path(sys.argv[0]).resolve().parent


def load_json_file(filename: str, fallback):
    p = app_dir() / filename
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return fallback
    return fallback


def classify_file(path: Path, include_unknown: bool) -> Optional[str]:
    for typ, pat in FILE_TYPE_PATTERNS:
        if pat.match(path.name):
            return typ
    for typ, pat in plugin_type_patterns():
        if pat.match(path.name):
            return typ
    if include_unknown and path.suffix.lower() in {".log", ".txt", ".out", ".ar"}:
        return "UNKNOWN"
    return None


def parse_ws_filename_datetime(path: Path) -> Optional[datetime]:
    m = WS_FILENAME_RE.match(path.name)
    if not m:
        return None
    try:
        return datetime(
            int(m.group("y")), MONTHS[m.group("mon").lower()], int(m.group("d")),
            int(m.group("h")), int(m.group("mi")), int(m.group("s"))
        )
    except Exception:
        return None


def parse_filename_datetime(path: Path) -> Optional[datetime]:
    """Extract a best-effort date/time from a log file name.

    Used when the log body has only a time value, as seen in MRServer and
    similar logs. If the file name contains only a date, midnight is returned;
    the row-level time is then combined with this date.
    """
    ws_dt = parse_ws_filename_datetime(path)
    if ws_dt:
        return ws_dt

    name = path.name
    patterns: list[re.Pattern[str]] = [
        # MRServer example: mrserver_Thu_Jul__2_17_02_10_2026.log
        # Use the date/time embedded in the file name only as the starting date.
        re.compile(r"(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)[_ -]+(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[_ -]+(?P<d>\d{1,2})[_ -]+(?P<h>\d{1,2})[_ -]+(?P<mi>\d{1,2})(?:[_ -]+(?P<s>\d{1,2}))?[_ -]+(?P<y>20\d{2})", re.I),
        re.compile(r"(?P<y>20\d{2})[_-](?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[_-](?P<d>\d{1,2})(?:[_-](?P<h>\d{1,2})[_-](?P<mi>\d{1,2})(?:[_-](?P<s>\d{1,2}))?)?", re.I),
        re.compile(r"(?P<y>20\d{2})[-_](?P<m>\d{1,2})[-_](?P<d>\d{1,2})(?:[_ T-](?P<h>\d{1,2})[-_:](?P<mi>\d{1,2})(?:[-_:](?P<s>\d{1,2}))?)?", re.I),
        re.compile(r"(?P<y>20\d{2})(?P<m>\d{2})(?P<d>\d{2})(?:[_-]?(?P<h>\d{2})(?P<mi>\d{2})(?P<s>\d{2})?)?", re.I),
        re.compile(r"(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[_-](?P<d>\d{1,2})[_-](?P<y>20\d{2})(?:[_-](?P<h>\d{1,2})[_-](?P<mi>\d{1,2})(?:[_-](?P<s>\d{1,2}))?)?", re.I),
    ]
    for pat in patterns:
        m = pat.search(name)
        if not m:
            continue
        try:
            gd = m.groupdict()
            year = int(gd["y"])
            month = MONTHS[gd["mon"].lower()] if gd.get("mon") else int(gd["m"])
            day = int(gd["d"])
            hour = int(gd.get("h") or 0)
            minute = int(gd.get("mi") or 0)
            second = int(gd.get("s") or 0)
            return datetime(year, month, day, hour, minute, second)
        except Exception:
            continue
    return None


def _microsecond_from_match(gd: dict[str, str]) -> int:
    ms = gd.get("ms") or ""
    if not ms:
        return 0
    return int((ms + "000000")[:6])


def _datetime_from_match(gd: dict[str, str]) -> datetime:
    year = int(gd["y"])
    month = MONTHS[gd["mon"].lower()] if gd.get("mon") else int(gd["m"])
    day = int(gd["d"])
    hour = int(gd.get("h") or 0)
    minute = int(gd.get("mi") or 0)
    second = int(gd.get("s") or 0)
    microsecond = _microsecond_from_match(gd)
    return datetime(year, month, day, hour, minute, second, microsecond)




GESYS_SECTION_DATETIME_RE = re.compile(
    r"^\s*\d+\s+\d+\s+\d+\s+(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+"
    r"(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+"
    r"(?P<d>\d{1,2})\s+"
    r"(?P<h>\d{1,2}):(?P<mi>\d{2}):(?P<s>\d{2})\s+"
    r"(?P<y>20\d{2})\b",
    re.I,
)

LAIS_DATE_HEADER_RE = re.compile(
    r"^\s*(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+"
    r"(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+"
    r"(?P<d>\d{1,2})\s+"
    r"(?P<y>20\d{2})\s*$",
    re.I,
)


def parse_lais_date_header(text: str) -> Optional[date]:
    """Parse LAIS section date headers such as 'Mon Jan 26 2026'.

    LAIS logs are date-section based: a standalone weekday/month/day/year line
    defines the date for following time-led rows.  The following rows start with
    high precision times such as '16:44:32.390869'.  File date and unrelated
    embedded text must not override these section dates.
    """
    m = LAIS_DATE_HEADER_RE.match(text)
    if not m:
        return None
    try:
        gd = m.groupdict()
        return date(int(gd["y"]), MONTHS[gd["mon"].lower()], int(gd["d"]))
    except Exception:
        return None


def parse_gesys_section_datetime(text: str) -> Optional[datetime]:
    """Parse the timestamp line that starts a GESYS log section.

    GESYS records are section-based. A section timestamp appears on a line like:
        1780144501    0    1    Sat May 30 12:35:01 2026    0
    Dates inside these section headers are the authoritative record timestamp.
    File name dates and old release/build dates must not be used for GESYS rows.
    """
    m = GESYS_SECTION_DATETIME_RE.search(text)
    if not m:
        return None
    try:
        gd = m.groupdict()
        return datetime(
            int(gd["y"]),
            MONTHS[gd["mon"].lower()],
            int(gd["d"]),
            int(gd["h"]),
            int(gd["mi"]),
            int(gd["s"]),
        )
    except Exception:
        return None


def parse_datetime_from_text(text: str, fallback_date: Optional[date] = None) -> Optional[datetime]:
    """Return a full timestamp from log content.

    Priority is always the timestamp written inside the log text.  The optional
    fallback_date is used only for time-only lines, mainly WS logs whose date is
    stored in the file name but whose record timestamp is at the line head.
    """
    for pat in FULL_DATETIME_PATTERNS:
        m = pat.search(text)
        if not m:
            continue
        try:
            return _datetime_from_match(m.groupdict())
        except Exception:
            continue

    if fallback_date:
        m = TIME_AT_LINE_START_RE.search(text)
        if m:
            try:
                gd = m.groupdict()
                return datetime(
                    fallback_date.year, fallback_date.month, fallback_date.day,
                    int(gd["h"]), int(gd["mi"]), int(gd["s"]), _microsecond_from_match(gd)
                )
            except Exception:
                return None
    return None


def parse_date_only_from_text(text: str) -> Optional[date]:
    for pat in DATE_ONLY_PATTERNS:
        m = pat.search(text)
        if not m:
            continue
        try:
            gd = m.groupdict()
            year = int(gd["y"])
            month = MONTHS[gd["mon"].lower()] if gd.get("mon") else int(gd["m"])
            day = int(gd["d"])
            return date(year, month, day)
        except Exception:
            continue
    return None


def parse_time_at_line_start(text: str, current_date: Optional[date], previous_ts: Optional[datetime] = None) -> Optional[datetime]:
    if current_date is None:
        return None
    m = TIME_AT_LINE_START_RE.search(text)
    if not m:
        return None
    gd = m.groupdict()
    try:
        candidate = datetime(
            current_date.year, current_date.month, current_date.day,
            int(gd["h"]), int(gd["mi"]), int(gd["s"]), _microsecond_from_match(gd)
        )
        # Filename-date fallback logs such as MRServer often contain time-only
        # records. The file is read from top to bottom as chronological order.
        # When the record time goes backwards, treat it as a new day.
        # This intentionally does NOT use file created/modified time.
        if previous_ts and candidate < previous_ts:
            while candidate < previous_ts:
                candidate += timedelta(days=1)
        return candidate
    except Exception:
        return None



def _trusted_date_position(text: str, start_index: int) -> bool:
    """Accept dates only when they look like the record date/time, not a version date."""
    prefix = text[:start_index].strip().lower()
    if start_index <= 3:
        return True
    if prefix.endswith(("date:", "time:", "timestamp:", "date", "time", "timestamp")):
        return True
    return False


def parse_trusted_full_datetime_from_text(text: str) -> Optional[datetime]:
    for pat in FULL_DATETIME_PATTERNS:
        m = pat.search(text)
        if not m or not _trusted_date_position(text, m.start()):
            continue
        try:
            return _datetime_from_match(m.groupdict())
        except Exception:
            continue
    return None

def is_explicit_date_header(line: str) -> bool:
    """Return True only when a line is intended to define the log date.

    Do not treat dates embedded in ordinary messages such as
    "Neuro Release Reg _6.33 [v Jun 14 2022 ...]" as the record date.
    CSA/CGA/MRServer logs often have a reliable time at the beginning of each
    row and old version/build dates in the message body.
    """
    t = line.strip()
    low = t.lower()
    if not t:
        return False
    if low.startswith(("date:", "log date:", "current date:", "system date:", "start date:")):
        return True
    if re.match(r"^(date|log date|current date|system date|start date)\s*=", low):
        return True
    # Allow a line that is only a date, optionally with separators.
    if re.match(r"^(?:(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+)?(?:20\d{2}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}/\d{1,2}/20\d{2}|\d{1,2}[- ](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ]20\d{2}|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/ ]\d{1,2}[-/ ,]+20\d{2})\s*$", t, re.I):
        return True
    return False


def extract_content_timestamp(line: str, source_type: str, current_date: Optional[date], previous_ts: Optional[datetime]) -> tuple[Optional[datetime], Optional[date]]:
    """Extract the record timestamp from log content.

    v28 rule:
    - For time-led logs such as WS/CSA/CGA/MRServer, the time at the very
      beginning of the row is authoritative. It is combined with the current
      date, which usually comes from the file name.
    - Full date/time inside the message is used only when no row-leading time
      exists. This prevents old software/version dates in CSA/CGA messages from
      overwriting the real record timestamp.
    - Date-only text changes current_date only when it is an explicit date
      header, not an arbitrary date in the message body.
    """
    source_upper = str(source_type).upper()

    # LAIS is date-section based. A standalone line such as
    # "Mon Jan 26 2026" defines the date for following time-led rows.
    # Do not use file name dates for LAIS when the section date is present.
    if source_upper == "LAIS":
        lais_date = parse_lais_date_header(line)
        if lais_date:
            return None, lais_date

    # GESYS is section-based. Use the section timestamp line first and let
    # following lines inherit it until the next section. This prevents the
    # filename date or unrelated dates from being applied to all GESYS rows.
    if source_upper == "GESYS":
        gesys_ts = parse_gesys_section_datetime(line)
        if gesys_ts:
            return gesys_ts, gesys_ts.date()

    # First priority for time-led logs: row-leading time + current/file date.
    line_ts = parse_time_at_line_start(line, current_date, previous_ts)
    if line_ts:
        return line_ts, line_ts.date()

    # If there is no leading time, use a full date/time only when it is a
    # trusted record timestamp/header. Do not use embedded release/build dates.
    full_ts = parse_trusted_full_datetime_from_text(line)
    if full_ts:
        return full_ts, full_ts.date()

    if is_explicit_date_header(line):
        date_header = parse_date_only_from_text(line)
        if date_header:
            current_date = date_header

    return None, current_date


def parse_user_date(value: Optional[str]) -> Optional[date]:
    """Parse GUI date text into a date.

    QDateEdit may return formats such as yyyy/MM/dd, yyyy-MM-dd, or
    locale-specific strings.  This helper is intentionally small and returns a
    date object because the date range filter is inclusive by day.
    """
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%m/%d/%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass
    dt = parse_datetime_from_text(text)
    if dt:
        return dt.date()
    d = parse_date_only_from_text(text)
    return d

def in_date_range(ts: Optional[datetime], file_date: Optional[date], start: Optional[date], end: Optional[date]) -> bool:
    # v8: date filtering is based on the timestamp extracted from the log body.
    # file_date is kept only for backward call compatibility and is not used.
    if start is None and end is None:
        return True
    if ts is None:
        return False
    d = ts.date()
    if start and d < start:
        return False
    if end and d > end:
        return False
    return True

def detect_level(line: str) -> str:
    m = LEVEL_RE.search(line)
    if not m:
        return ""
    return m.group(1).upper()


def detect_category(source_type: str, line: str) -> str:
    u = line.upper()
    if "ERROR" in u or "FAIL" in u or "FATAL" in u or "ALARM" in u:
        return "ERROR"
    if "WARN" in u:
        return "WARNING"
    if source_type == "WS":
        if "CIRCULATE" in u:
            return "CIRCULATE"
        if "DRAIN" in u:
            return "DRAIN"
        if "FILL" in u:
            return "FILL"
        if "PAUSE" in u:
            return "PAUSE"
    return source_type


def clean_excel_value(value):
    """Return a value safe for openpyxl worksheet cells.

    Field logs sometimes contain control characters such as SUB/ESC. Excel
    cannot store those characters in xlsx XML and openpyxl raises
    IllegalCharacterError during save. Keep the readable text and remove only
    characters that are illegal in worksheets.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        try:
            value = ILLEGAL_CHARACTERS_RE.sub("", value)
        except Exception:
            value = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", value)
        return value.replace("\x00", "")
    return value


def clean_message(line: str) -> str:
    return str(clean_excel_value(line)).strip()


def viewer_message_body(text: str) -> str:
    """Message column text for Explorer only.

    Keep the raw record in the detail pane, but do not repeat the leading
    timestamp in the visible Message column.
    """
    t = str(clean_excel_value(text or "")).strip()
    t = TIME_AT_LINE_START_RE.sub("", t, count=1).strip()
    return t


def format_timestamp(ts: Optional[datetime]) -> str:
    """Format every output timestamp explicitly as yyyy/mm/dd hh:mm:ss.f.

    Missing date/time components are zero-padded by datetime/strftime.
    If the source line has no fractional seconds, append .0 so every row has
    an explicit timestamp like 2026/07/04 13:07:58.6.
    """
    if not ts:
        return ""
    base = ts.strftime("%Y/%m/%d %H:%M:%S")
    if ts.microsecond:
        frac = f"{ts.microsecond:06d}".rstrip("0")
        return f"{base}.{frac}"
    return f"{base}.0"


def record_output_headers() -> list[str]:
    return ["Timestamp", "SourceType", "File", "Line", "Level", "Category", "Message", "Raw"]


def record_to_output_row(r: "LogRecord") -> list[Any]:
    # Default: write Timestamp as Excel datetime. Writer applies display format.
    return [
        r.timestamp if r.timestamp else "",
        r.source_type, r.filename, r.line_no, r.level, r.category, r.message, r.raw
    ]


def record_to_output_row_fast(r: "LogRecord", timestamp_as_text: bool = False) -> list[Any]:
    ts_value = format_timestamp(r.timestamp) if timestamp_as_text else (r.timestamp if r.timestamp else "")
    return [ts_value, r.source_type, r.filename, r.line_no, r.level, r.category, r.message, r.raw]


def csa_hit_to_output_row(h: dict[str, Any]) -> list[Any]:
    ts = h.get("Timestamp")
    return [
        ts if isinstance(ts, datetime) else "",
        h.get("File", ""), h.get("Line", ""), h.get("Keyword", ""), h.get("Message", ""), h.get("Raw", "")
    ]


def read_text_lines(path: Path) -> list[str]:
    encodings = ("utf-8-sig", "utf-8", "cp932", "shift_jis", "latin-1")
    last_exc: Optional[Exception] = None
    for enc in encodings:
        try:
            return path.read_text(encoding=enc, errors="strict").splitlines()
        except Exception as exc:
            last_exc = exc
    try:
        return path.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        raise RuntimeError(f"Unable to read {path}: {last_exc}")


def parse_file(path: Path, source_type: str) -> list[LogRecord]:
    # v14 timestamp rule:
    # 1) Full date+time written in the log body is used first.
    # 2) If the body has only a time, combine it with the body date header or,
    #    when no body date exists, the date parsed from the file name.
    #    File created/modified dates are intentionally never used.
    # 3) Within one file, lines are read from top to bottom. If the line time
    #    rolls back across midnight, the date is advanced by one day.
    # 4) Lines with no timestamp inherit the previous row timestamp; if there is
    #    no previous row timestamp, the file-name datetime is used as fallback.
    fallback_dt = parse_filename_datetime(path)
    current_date: Optional[date] = fallback_dt.date() if fallback_dt else None
    records: list[LogRecord] = []
    lines = read_text_lines(path)
    last_ts: Optional[datetime] = None
    for idx, line in enumerate(lines, start=1):
        msg = clean_message(line)
        if not msg:
            continue
        ts, current_date = extract_content_timestamp(msg, source_type, current_date, last_ts)
        if ts:
            last_ts = ts
        else:
            # Multi-line continuations inherit the last timestamp. If even the
            # first line has no time, fall back to the filename date/time so the
            # row still has an explicit timestamp when possible.
            ts = last_ts or fallback_dt
            if ts:
                last_ts = ts
        level = detect_level(msg)
        category = detect_category(source_type, msg)
        records.append(LogRecord(ts, source_type, path.name, idx, level, category, msg, line.rstrip("\r\n")))
    if not records and fallback_dt:
        records.append(LogRecord(fallback_dt, source_type, path.name, 0, "", source_type, "File detected but no readable lines", ""))
    return records

def iter_files(root: Path, recursive: bool) -> Iterable[Path]:
    it = root.rglob("*") if recursive else root.glob("*")
    for p in it:
        if p.is_file():
            yield p


def enabled_types(options: RunOptions) -> set[str]:
    enabled = set()
    if options.include_ws: enabled.add("WS")
    if options.include_cga: enabled.add("CGA")
    if options.include_csa: enabled.add("CSA")
    if options.include_mrserver: enabled.add("MRSERVER")
    if options.include_gesys: enabled.add("GESYS")
    if options.include_lais: enabled.add("LAIS")
    if options.include_psc: enabled.add("PSC")
    if options.include_unknown: enabled.add("UNKNOWN")
    for t in (getattr(options, "plugin_types", None) or []):
        if t:
            enabled.add(str(t))
    return enabled



PSC_DATE_HEADER_RE = re.compile(r"^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2}\s+\d{4}$", re.I)
PSC_TIME_RE = re.compile(r"^\d{2}:\d{2}:\d{2}(?:\.\d+)?$")
PSC_PARAM_RE = re.compile(r"([A-Za-z_][A-Za-z0-9_\[\]\.]*)(\s*=\s*)(\"[^\"]*\"|'[^']*'|[^\s;]*)")


def normalize_psc_date_header(line: str) -> str:
    parts = line.split()
    if len(parts) >= 4:
        try:
            return f"{int(parts[3]):04d}-{MONTHS[parts[1].lower()]:02d}-{int(parts[2]):02d}"
        except Exception:
            return line.strip()
    return line.strip()


def trim_trailing_punctuation(v: str) -> str:
    v = v.strip()
    while len(v) > 1 and v.endswith('.'):
        try:
            float(v[:-1])
            v = v[:-1]
        except Exception:
            break
    return v


def extract_psc_params(text: str) -> list[tuple[str, str]]:
    return [(m.group(1), trim_trailing_punctuation(m.group(3))) for m in PSC_PARAM_RE.finditer(text)]


def first_token(text: str) -> str:
    return text.strip().split()[0] if text.strip().split() else ""


def second_token(text: str) -> str:
    parts = text.strip().split()
    return parts[1] if len(parts) >= 2 else ""


def parse_psc_line(raw_line: str, current_date: str, source_file: str) -> Optional[dict[str, Any]]:
    raw_line = raw_line.strip()
    if not raw_line:
        return None
    first_space = raw_line.find(" ")
    if first_space <= 0:
        return None
    log_time = raw_line[:first_space]
    if not PSC_TIME_RE.match(log_time):
        return None
    rest = raw_line[first_space + 1:].strip()
    process_name = ""
    if rest.startswith("[") and "]" in rest:
        close = rest.find("]")
        process_name = rest[1:close]
        rest = rest[close + 1:].strip()
    full_dt = f"{current_date} {log_time}" if current_date else log_time
    module_name = ""
    function_name = ""
    raw_params = ""
    message_text = rest
    if "::" in rest:
        before, after = rest.split("::", 1)
        module_name = before.strip()
        if ":" in after:
            function_name, raw_params = after.split(":", 1)
            function_name = function_name.strip()
            raw_params = raw_params.strip()
        else:
            function_name = first_token(after)
            raw_params = after[len(function_name):].strip()
    else:
        if ":" in rest:
            module_name, temp = rest.split(":", 1)
            module_name = module_name.strip()
            temp = temp.strip()
            function_name = first_token(temp)
            raw_params = temp
        else:
            module_name = first_token(rest)
            function_name = second_token(rest)
            raw_params = rest
    params = extract_psc_params(rest)
    return {
        "SourceFile": source_file,
        "LogDate": current_date,
        "Time": log_time,
        "DateTime": full_dt,
        "Process": process_name,
        "Module": module_name,
        "Function": function_name.strip(),
        "Message": message_text,
        "ParameterText": raw_params,
        "ParamCount": len(params),
        "RawLine": raw_line,
        "Params": params,
    }


def parse_psc_file_detail(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[LogRecord]]:
    current_date = ""
    log_rows: list[dict[str, Any]] = []
    param_rows: list[dict[str, Any]] = []
    records: list[LogRecord] = []
    max_param_count = 0
    for idx, line in enumerate(read_text_lines(path), start=1):
        raw = line.strip()
        if not raw:
            continue
        if PSC_DATE_HEADER_RE.match(raw):
            current_date = normalize_psc_date_header(raw)
            continue
        row = parse_psc_line(raw, current_date, path.name)
        if not row:
            continue
        params = row.pop("Params")
        max_param_count = max(max_param_count, len(params))
        for i, (name, value) in enumerate(params, start=1):
            row[f"Param{i}"] = f"{name}={value}"
            param_rows.append({
                "SourceFile": path.name,
                "LogDate": row["LogDate"],
                "Time": row["Time"],
                "DateTime": row["DateTime"],
                "Process": row["Process"],
                "Module": row["Module"],
                "Function": row["Function"],
                "ParameterName": name,
                "ParameterValue": value,
                "Message": row["Message"],
                "RawLine": row["RawLine"],
            })
        log_rows.append(row)
        ts = parse_datetime_from_text(str(row.get("DateTime", "")))
        records.append(LogRecord(ts, "PSC", path.name, idx, detect_level(raw), detect_category("PSC", raw), row.get("Message", raw), raw))
    return log_rows, param_rows, records


REVIEW_KEY_RE = re.compile(r"(^|[\t]+| {2,})([^:\t]{1,90}?)\s*:")


def clean_control_text(s: str) -> str:
    s = s.replace("\u00a0", " ").replace("\t", " ")
    s = "".join(ch for ch in s if ord(ch) >= 32)
    while "  " in s:
        s = s.replace("  ", " ")
    return s.strip()


def clean_review_value(s: str) -> str:
    s = s.replace("\u00a0", " ").replace("\t", " ")
    s = "".join(ch for ch in s if ord(ch) >= 32)
    return s.strip()


def is_review_section_header(s: str) -> bool:
    t = s.strip(); u = t.upper()
    if not t or ":" in t or t.startswith("-----"):
        return False
    return (u.endswith("SCREEN") or " SCREEN" in u or u in {"GENERAL INFORMATION", "USER/SYSTEM PREFERENCE STATE", "ACCELERATION", "USER CV", "USER CV SCREEN"})


def review_value_after_colon(s: str) -> str:
    return clean_review_value(s.split(":", 1)[1]) if ":" in s else ""


def normalize_review_key(s: str) -> str:
    t = clean_control_text(s)
    if t.lower().endswith(" is"):
        t = clean_control_text(t[:-3])
    if t.upper() == "DATE":
        return "date"
    if t.upper() == "TIME":
        return "time"
    if t.upper() == "ID":
        return "ID"
    return t


def extract_review_key_values(raw_line: str, section_name: str, scan_data: dict[str, Any]) -> None:
    matches = list(REVIEW_KEY_RE.finditer(raw_line))
    if not matches:
        return
    for i, m in enumerate(matches):
        raw_key = normalize_review_key(m.group(2))
        value_start = m.end()
        value_end = matches[i + 1].start() if i < len(matches) - 1 else len(raw_line)
        raw_value = clean_review_value(raw_line[value_start:value_end])
        if not raw_key:
            continue
        if raw_key in {"date", "time", "ID"}:
            scan_data[raw_key] = raw_value
        else:
            full_key = f"{clean_control_text(section_name)} - {raw_key}" if section_name else raw_key
            scan_data[full_key] = raw_value


def parse_reviewout_file(path: Path) -> list[dict[str, Any]]:
    scan_list: list[dict[str, Any]] = []
    scan_data: dict[str, Any] = {}
    section_name = ""
    for line in read_text_lines(path):
        raw_line = line.replace("\u00a0", " ").replace("\r", "")
        trimmed = raw_line.strip()
        if not trimmed:
            continue
        low = trimmed.lower()
        if low.startswith("date:"):
            if scan_data:
                scan_list.append(scan_data)
                scan_data = {}
            section_name = ""
            scan_data["date"] = review_value_after_colon(trimmed)
        elif low.startswith("time:"):
            scan_data["time"] = review_value_after_colon(trimmed)
        elif low.startswith("id") and ":" in low:
            scan_data["ID"] = review_value_after_colon(trimmed)
        elif is_review_section_header(trimmed):
            section_name = clean_control_text(trimmed)
        elif ":" in trimmed:
            extract_review_key_values(raw_line, section_name, scan_data)
    if scan_data:
        scan_list.append(scan_data)
    return scan_list


def write_generic_table_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]], table_name: str, progress_cb=None, progress_start: int = 0, progress_total: int = 0, control=None) -> int:
    ws = wb.create_sheet(title)
    written = write_rows(ws, headers, rows, progress_cb=progress_cb, progress_start=progress_start, progress_total=progress_total, progress_label=f"Writing {title}", control=control)
    apply_table(ws, table_name)
    autosize(ws)
    return written


def import_psc_only(file_path: str, output_folder: str = "") -> str:
    path = Path(file_path)
    out_dir = Path(output_folder) if output_folder else path.parent
    log_rows, param_rows, _ = parse_psc_file_detail(path)
    if not log_rows:
        raise ValueError("No valid PSC rows were detected in the selected file.")
    wb = Workbook()
    ws = wb.active
    ws.title = "PSC_Log"
    max_param = max((int(r.get("ParamCount", 0)) for r in log_rows), default=0)
    log_headers = ["SourceFile", "LogDate", "Time", "DateTime", "Process", "Module", "Function", "Message", "ParameterText", "ParamCount", "RawLine"] + [f"Param{i}" for i in range(1, max_param + 1)]
    write_rows(ws, log_headers, [[r.get(h, "") for h in log_headers] for r in log_rows])
    apply_table(ws, "PSCLogTable")
    autosize(ws)
    param_headers = ["SourceFile", "LogDate", "Time", "DateTime", "Process", "Module", "Function", "ParameterName", "ParameterValue", "Message", "RawLine"]
    write_generic_table_sheet(wb, "PSC_Params", param_headers, [[r.get(h, "") for h in param_headers] for r in param_rows], "PSCParamTable")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"{path.stem}_PSC_Import_{stamp}.xlsx"
    wb.save(out_path)
    strip_excel_table_xml(out_path)
    return str(out_path)


def import_reviewout_only(file_path: str, output_folder: str = "") -> str:
    path = Path(file_path)
    out_dir = Path(output_folder) if output_folder else path.parent
    scan_list = parse_reviewout_file(path)
    if not scan_list:
        raise ValueError("No valid review.out scan records were detected in the selected file.")
    all_keys: list[str] = ["date", "time"]
    if any("ID" in item for item in scan_list):
        all_keys.append("ID")
    for item in scan_list:
        for k in item.keys():
            if k not in all_keys:
                all_keys.append(k)
    wb = Workbook()
    ws = wb.active
    ws.title = "ScanSummary"
    write_rows(ws, all_keys, [[item.get(h, "") for h in all_keys] for item in scan_list])
    apply_table(ws, "ReviewOutScanSummary")
    autosize(ws)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = path.name
    if "." in base:
        base = base.rsplit(".", 1)[0]
    if base.lower().endswith(".out"):
        base = base[:-4]
    out_path = out_dir / f"{base}_reviewout_Import_{stamp}.xlsx"
    wb.save(out_path)
    strip_excel_table_xml(out_path)
    return str(out_path)



def import_selected_file_only(file_path: str, output_folder: str = "") -> str:
    """Import one selected file without merge.

    This is intentionally a file-level import/export path. It does not merge
    with other logs and it applies the same timestamp parser/output writer as
    the main Search sheet so operators can inspect one file quickly.
    """
    path = Path(file_path)
    out_dir = Path(output_folder) if output_folder else path.parent
    if is_review_file(path.name):
        return import_reviewout_only(file_path, output_folder)
    typ = classify_file(path, True) or "UNKNOWN"
    if typ == "PSC":
        return import_psc_only(file_path, output_folder)
    records = parse_file(path, typ)
    if not records:
        raise ValueError("No readable rows were detected in the selected file.")
    records.sort(key=lambda r: (r.timestamp is None, r.timestamp or datetime.max, r.line_no))
    wb = Workbook()
    ws = wb.active
    ws.title = "Search"
    headers = record_output_headers()
    write_rows(ws, headers, [record_to_output_row(r) for r in records])
    apply_table(ws, "Search")
    autosize(ws)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100
    activate_search_sheet(wb, ws)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", path.stem)[:80]
    out_path = out_dir / f"{safe_stem}_{typ}_Import_{stamp}.xlsx"
    wb.save(out_path)
    strip_excel_table_xml(out_path)
    return str(out_path)

def check_csa_errors(records: list[LogRecord], rules: list[dict]) -> list[dict]:
    hits: list[dict] = []
    compiled = []
    for r in rules:
        kw = str(r.get("keyword", "")).strip()
        if not kw:
            continue
        compiled.append((kw, str(r.get("message", kw)), re.compile(re.escape(kw), re.I)))
    for rec in records:
        if rec.source_type not in {"CSA", "UNKNOWN"}:
            continue
        for kw, msg, pat in compiled:
            if pat.search(rec.raw) or pat.search(rec.message):
                hits.append({
                    "Timestamp": rec.timestamp,
                    "File": rec.filename,
                    "Line": rec.line_no,
                    "Keyword": kw,
                    "Message": msg,
                    "Raw": rec.raw,
                })
    return hits


def autosize(ws, max_width: int = 70, max_scan_rows: int = 300) -> None:
    """Fast autosize.

    Large log sheets can contain many rows. Scanning every cell for width makes
    the app look frozen after the file scan reaches 100%, so only the header and
    first max_scan_rows rows are sampled.
    """
    max_row = min(ws.max_row, max_scan_rows + 1)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for row in range(1, max_row + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            width = max(width, min(max_width, len(str(value)) + 2))
        ws.column_dimensions[letter].width = width


def apply_table(ws, name: str, max_table_rows: int = 50000) -> None:
    """Do not create Excel Table objects.

    Previous versions added openpyxl Table objects and AutoFilter together on
    multiple sheets. Excel repaired the generated workbook and removed
    /xl/tables/table*.xml. For field reliability, v13 uses worksheet
    AutoFilter only. This keeps filter drop-downs without creating table XML.
    """
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    return


def write_rows(ws, headers: list[str], rows: list[list], fast_threshold: int = 50000, progress_cb=None, progress_start: int = 0, progress_total: int = 0, progress_label: str = "Writing rows", control=None, styles: bool = True, timestamp_format: bool = True) -> int:
    """Write rows to a worksheet with row-level progress callbacks.

    openpyxl cannot report progress during the final ZIP compression/save phase,
    but it can report progress while rows are appended.  For large outputs this
    keeps the UI popup moving instead of staying at 0% until the save completes.

    Returns the number of data rows written.
    """
    headers = [clean_excel_value(h) for h in headers]
    ws.append(headers)
    row_count = len(rows)
    update_every = 1000 if row_count < 50000 else 5000
    total_for_cb = progress_total or max(1, row_count)
    for i, row in enumerate(rows, start=1):
        if control:
            control.check()
        ws.append([clean_excel_value(v) for v in row])
        if progress_cb and (i == 1 or i % update_every == 0 or i == row_count):
            progress_cb(min(progress_start + i, total_for_cb), total_for_cb, f"{progress_label}: {i}/{row_count} rows")

    if styles:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fast path for large sheets. Avoid styling every cell.
        if row_count <= fast_threshold:
            thin = Side(style="thin", color="D9E2F3")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for cell in ws[1]:
                cell.border = border
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Required output behavior: freeze at C2 and enable filters.
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Timestamp must be explicit on every row as yyyy/mm/dd hh:mm:ss.0.
    if headers and str(headers[0]).strip().lower() == "timestamp":
        ws.column_dimensions["A"].width = 24
        if timestamp_format:
            for cell in ws["A"][1:]:
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy/mm/dd hh:mm:ss.0"
    return row_count


def activate_search_sheet(wb: Workbook, ws) -> None:
    """Open workbook on Search sheet with C2 selected/frozen."""
    try:
        wb.active = wb.sheetnames.index(ws.title)
    except Exception:
        pass
    try:
        ws.freeze_panes = "C2"
        ws.sheet_view.selection[0].activeCell = "C2"
        ws.sheet_view.selection[0].sqref = "C2"
    except Exception:
        pass



def strip_excel_table_xml(xlsx_path: Path) -> None:
    """Remove Excel Table parts from an xlsx as a final safety pass.

    Field Excel reported repairs in /xl/tables/table*.xml when older builds
    produced ListObject/Table XML. v15 uses worksheet AutoFilter only, but this
    sanitizer guarantees the final workbook contains no xl/tables parts even if
    a legacy code path accidentally creates them.
    """
    try:
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            return
        tmp_path = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")
        ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        ns_rel = "http://schemas.openxmlformats.org/package/2006/relationships"
        ET.register_namespace('', ns_main)
        ET.register_namespace('', ns_rel)
        with zipfile.ZipFile(xlsx_path, 'r') as zin, zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                name = item.filename
                if name.startswith('xl/tables/'):
                    continue
                data = zin.read(name)
                if name == '[Content_Types].xml':
                    try:
                        root = ET.fromstring(data)
                        for elem in list(root):
                            if elem.attrib.get('PartName', '').startswith('/xl/tables/'):
                                root.remove(elem)
                        data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                    except Exception:
                        pass
                elif name.startswith('xl/worksheets/_rels/') and name.endswith('.rels'):
                    try:
                        root = ET.fromstring(data)
                        for elem in list(root):
                            if '/table' in elem.attrib.get('Type', '') or elem.attrib.get('Target', '').startswith('../tables/'):
                                root.remove(elem)
                        data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                    except Exception:
                        pass
                elif name.startswith('xl/worksheets/sheet') and name.endswith('.xml'):
                    try:
                        root = ET.fromstring(data)
                        table_parts = root.find(f'{{{ns_main}}}tableParts')
                        if table_parts is not None:
                            root.remove(table_parts)
                        data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                    except Exception:
                        pass
                zout.writestr(item, data)
        tmp_path.replace(xlsx_path)
    except Exception:
        # Do not block output creation if the safety pass itself fails.
        pass

def safe_csv_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return format_timestamp(value)
    return str(value).replace("\x00", "")


def write_records_csv(records: list[LogRecord], csv_path: Path, log_cb=None, progress_cb=None, progress_offset: int = 0, progress_total: int = 0, control=None) -> None:
    if log_cb:
        log_cb(f"Writing fast CSV: {csv_path.name}")
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    headers = record_output_headers()
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for i, r in enumerate(records, start=1):
            if control and i % 1000 == 0:
                control.check()
            if progress_cb and (i % 5000 == 0 or i == len(records)):
                progress_cb(progress_offset + i, progress_total or len(records), f"Writing {csv_path.name}: {i}/{len(records)} rows")
            writer.writerow(record_to_output_row(r))


def write_dict_rows_csv(rows: list[dict[str, Any]], csv_path: Path, log_cb=None, progress_cb=None, progress_offset: int = 0, progress_total: int = 0, control=None) -> None:
    if not rows:
        return
    if log_cb:
        log_cb(f"Writing fast CSV: {csv_path.name}")
    headers: list[str] = []
    for row in rows:
        for k in row.keys():
            if k not in headers:
                headers.append(k)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for i, row in enumerate(rows, start=1):
            if control and i % 1000 == 0:
                control.check()
            if progress_cb and (i % 5000 == 0 or i == len(rows)):
                progress_cb(progress_offset + i, progress_total or len(rows), f"Writing {csv_path.name}: {i}/{len(rows)} rows")
            writer.writerow([safe_csv_value(row.get(h, "")) for h in headers])




def safe_sha256(path: Path, max_bytes: int = 1024 * 1024 * 64) -> str:
    """Return a SHA256 hash for traceability. Large files are still hashed in chunks."""
    import hashlib
    try:
        h = hashlib.sha256()
        with path.open("rb") as f:
            while True:
                b = f.read(1024 * 1024)
                if not b:
                    break
                h.update(b)
        return h.hexdigest()
    except Exception:
        return ""


def make_manifest_path(output_path: Path) -> Path:
    return output_path.with_name(output_path.stem + "_manifest.json")


def write_merge_manifest(output_path: Path, options: RunOptions, result: RunResult, source_entries: list[dict[str, Any]], extra_outputs: Optional[list[str]] = None, log_cb=None) -> Path:
    """Write a JSON sidecar describing exactly which files created a merge output."""
    manifest_path = make_manifest_path(output_path)
    data = {
        "schema": "LogMergeTool.MergeManifest.v1",
        "tool_version": APP_VERSION,
        "created_at": datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
        "output_file": str(output_path),
        "output_name": output_path.name,
        "mode": "merge",
        "site": options.site,
        "serial": options.serial,
        "source_folder": options.source_folder,
        "output_folder": options.output_folder,
        "recursive": bool(options.recursive),
        "date_filter": {
            "enabled": bool(options.use_start or options.use_end),
            "start": options.start_date if options.use_start else "",
            "end": options.end_date if options.use_end else "",
            "basis": "content timestamp per row; filename date fallback only when body date is unavailable; file created/modified date is never used",
        },
        "performance": {
            "turbo_csv": bool(getattr(options, "turbo_csv", False)),
            "summary_xlsx_only": bool(getattr(options, "summary_xlsx_only", False)),
            "fast_xlsx_mode": bool(getattr(options, "fast_xlsx_mode", False)),
            "csv_summary_mode": bool(getattr(options, "csv_summary_mode", False)),
            "skip_merged1": bool(getattr(options, "skip_merged1", False)),
            "timestamp_as_text": bool(getattr(options, "timestamp_as_text", False)),
            "disable_styles_large": bool(getattr(options, "disable_styles_large", False)),
            "worker_count": int(getattr(options, "worker_count", 0) or 0),
        },
        "noise_rules": {
            "enabled_in_merge": bool(getattr(options, "noise_enabled_merge", False)),
            "exclude_output": bool(getattr(options, "noise_exclude_output", False)),
            "learning": bool(getattr(options, "noise_learning", False)),
        },
        "run_result": asdict(result),
        "source_files": source_entries,
        "extra_outputs": extra_outputs or [],
        "split_support": {
            "supported": True,
            "note": "Split uses Search sheet or Search CSV rows grouped by SourceType and File. It reconstructs filtered output rows, not records removed by Noise rules.",
        },
    }
    manifest_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    if log_cb:
        log_cb(f"Manifest written: {manifest_path}")
    return manifest_path


def discover_related_manifest(merge_path: Path) -> Optional[Path]:
    candidates = [
        merge_path.with_name(merge_path.stem + "_manifest.json"),
        merge_path.with_suffix(".json"),
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def split_merge_project(merge_file: str, output_folder: str, progress_cb=None, log_cb=None) -> str:
    """Split a merged Search sheet/CSV back into per-source reconstructed logs.

    This reconstructs what exists in the merged output. If Noise rules excluded rows
    before output, those excluded rows cannot be reconstructed from the merge file.
    """
    def log(msg: str):
        if log_cb:
            log_cb(msg)
    src = Path(merge_file)
    if not src.exists():
        raise FileNotFoundError(str(src))
    out_root = Path(output_folder) if output_folder else src.parent / (src.stem + "_split")
    out_root.mkdir(parents=True, exist_ok=True)
    rows_by_key: dict[tuple[str, str], list[str]] = {}
    count = 0

    def add_row(row: dict[str, Any]):
        nonlocal count
        typ = str(row.get("SourceType") or row.get("Type") or "UNKNOWN").strip() or "UNKNOWN"
        fname = str(row.get("File") or row.get("Filename") or f"{typ}.log").strip() or f"{typ}.log"
        raw = str(row.get("Raw") or row.get("Message") or "")
        rows_by_key.setdefault((typ, fname), []).append(raw)
        count += 1
        if progress_cb and count % 5000 == 0:
            progress_cb(count, max(count, count + 1), f"Reading merged rows: {count}")

    if src.suffix.lower() == ".csv":
        with src.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                add_row(row)
    else:
        wb = load_workbook(src, read_only=True, data_only=True)
        sheet_name = "Search" if "Search" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        headers = [str(x or "").strip() for x in next(it)]
        for vals in it:
            row = {h: vals[i] if i < len(vals) else "" for i, h in enumerate(headers)}
            add_row(row)
        wb.close()

    written_files = []
    total_groups = max(1, len(rows_by_key))
    for idx, ((typ, fname), lines) in enumerate(sorted(rows_by_key.items()), start=1):
        safe_type = re.sub(r"[^A-Za-z0-9_-]+", "_", typ) or "UNKNOWN"
        safe_name = Path(re.sub(r"[\\/:*?\"<>|]+", "_", fname)).name or f"{safe_type}.log"
        target_dir = out_root / safe_type
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / safe_name
        # If multiple groups would collide, add an index suffix.
        if target.exists():
            target = target_dir / f"{target.stem}_{idx}{target.suffix or '.log'}"
        target.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8", errors="replace")
        written_files.append({"type": typ, "file": fname, "output": str(target), "rows": len(lines)})
        if progress_cb:
            progress_cb(idx, total_groups, f"Writing {target.name}")

    manifest = discover_related_manifest(src)
    split_manifest = {
        "schema": "LogMergeTool.SplitManifest.v1",
        "tool_version": APP_VERSION,
        "created_at": datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
        "source_merge_file": str(src),
        "source_manifest": str(manifest) if manifest else "",
        "output_folder": str(out_root),
        "total_rows": count,
        "files_written": written_files,
        "note": "This split recreates merged/filtered output rows. Rows not present in the merge output cannot be recovered.",
    }
    (out_root / "split_manifest.json").write_text(json.dumps(split_manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"Split completed: {out_root}")
    return str(out_root)

def build_summary_only_xlsx(records: list[LogRecord], csa_hits: list[dict], result: RunResult, options: RunOptions, output_path: Path, csv_files: list[str], log_cb=None) -> None:
    def outlog(msg: str):
        if log_cb:
            log_cb(msg)
    outlog("Creating summary xlsx workbook...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"] = "Log Merge Tool - No Excel / v13 Timestamp + C2 Freeze + Filter (No Tables)"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    summary_rows = [
        ["Version", APP_VERSION],
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Output Mode", "Turbo CSV + Summary xlsx"],
        ["Source Folder", options.source_folder],
        ["Output Folder", options.output_folder],
        ["Recursive", "Yes" if options.recursive else "No"],
        ["Start Date", options.start_date if options.use_start else "Not specified"],
        ["End Date", options.end_date if options.use_end else "Not specified"],
        ["Serial Number", options.serial],
        ["Site Name", options.site],
        ["Scanned Files", result.scanned_files],
        ["Parsed Files", result.parsed_files],
        ["Skipped Files", result.skipped_files],
        ["Total Records", result.total_records],
        ["Included Records", result.included_records],
        ["CSA Error Count", result.csa_error_count],
    ]
    for r, row in enumerate(summary_rows, start=3):
        ws.cell(r, 1, row[0]).font = Font(bold=True)
        ws.cell(r, 2, row[1])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 90

    def write_summary_sheet_rows(ws, headers, rows, label: str = "Writing sheet"):
        """Summary workbook writer helper.

        v24/v25 could call a local function named write_sheet_rows that only
        existed in the full-xlsx writer path.  When CSV + Summary xlsx mode was
        selected, that caused NameError during summary creation.  Keep this
        helper local and use the shared safe write_rows implementation.
        """
        outlog(label)
        return write_rows(ws, headers, rows, styles=True, timestamp_format=True)

    ws_files = wb.create_sheet("Output Files")
    write_summary_sheet_rows(ws_files, ["File", "Note"], [[f, "Open with Excel. Full merged data is here." if f.lower().endswith('.csv') else "Workbook output"] for f in csv_files], "Writing Output Files sheet")
    autosize(ws_files)

    type_counts: dict[str, int] = {}
    for r in records:
        type_counts[r.source_type] = type_counts.get(r.source_type, 0) + 1
    ws_type = wb.create_sheet("Type_Summary")
    write_summary_sheet_rows(ws_type, ["SourceType", "RecordCount"], [[k, type_counts[k]] for k in sorted(type_counts.keys())], "Writing Type Summary")
    autosize(ws_type)

    error_records = [r for r in records if r.category in {"ERROR", "WARNING"} or r.level in {"ERROR", "ERR", "WARN", "WARNING", "FATAL", "ALARM"}]
    ws_errors = wb.create_sheet("Errors_Sample")
    headers = record_output_headers()
    sample = error_records[:5000]
    write_summary_sheet_rows(ws_errors, headers, [record_to_output_row(r) for r in sample], "Writing Error Sample")
    autosize(ws_errors)

    ws_csa_hits = wb.create_sheet("CSA Error Hits")
    csa_headers = ["Timestamp", "File", "Line", "Keyword", "Message", "Raw"]
    csa_rows = [csa_hit_to_output_row(h) for h in csa_hits[:5000]]
    write_summary_sheet_rows(ws_csa_hits, csa_headers, csa_rows, "Writing CSA Error Hits")
    autosize(ws_csa_hits)

    ws_rules = wb.create_sheet("CSA Error List")
    write_summary_sheet_rows(ws_rules, ["Keyword", "Message"], [[r.get("keyword", ""), r.get("message", "")] for r in load_json_file("csa_error_rules.json", DEFAULT_CSA_RULES)], "Writing CSA Error List")
    autosize(ws_rules)

    ws_lists = wb.create_sheet("Lists")
    write_summary_sheet_rows(ws_lists, ["Serial Number", "Site Name"], [[r.get("serial", ""), r.get("site", "")] for r in load_json_file("site_serial_map.json", DEFAULT_SITE_MAP)], "Writing Lists sheet")
    autosize(ws_lists)

    outlog("Saving summary xlsx file...")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    strip_excel_table_xml(output_path)
    outlog("summary xlsx save completed.")

def build_output(records: list[LogRecord], csa_hits: list[dict], result: RunResult, options: RunOptions, output_path: Path, psc_log_rows: Optional[list[dict[str, Any]]] = None, psc_param_rows: Optional[list[dict[str, Any]]] = None, log_cb=None, progress_cb=None, control=None) -> None:
    def outlog(msg: str):
        if log_cb:
            log_cb(msg)
    # v24 Fast Excel Writer. For very large outputs, writing full xlsx is slow
    # because the final ZIP/save phase has no internal progress. CSV + summary
    # xlsx is therefore available as a practical high-speed path.
    turbo_threshold = 100000
    if getattr(options, "turbo_csv", True) and getattr(options, "summary_xlsx_only", True) and getattr(options, "csv_summary_mode", True) and len(records) >= turbo_threshold:
        outlog(f"Turbo CSV mode enabled for {len(records)} rows.")
        if control:
            control.check()
        sorted_records = sorted(records, key=lambda r: r.timestamp or datetime.min, reverse=True)
        base = output_path.with_suffix("")
        csv_files: list[str] = []
        search_csv = base.parent / f"{base.name}_Search.csv"
        write_total = max(1, len(sorted_records) + len(psc_log_rows or []) + len(psc_param_rows or []))
        written_offset = 0
        write_records_csv(sorted_records, search_csv, log_cb, progress_cb, written_offset, write_total, control)
        written_offset += len(sorted_records)
        csv_files.append(str(search_csv))
        error_records = [r for r in sorted_records if r.category in {"ERROR", "WARNING"} or r.level in {"ERROR", "ERR", "WARN", "WARNING", "FATAL", "ALARM"}]
        if error_records:
            error_csv = base.parent / f"{base.name}_Errors.csv"
            write_records_csv(error_records, error_csv, log_cb, None, 0, 0, control)
            csv_files.append(str(error_csv))
        if psc_log_rows:
            psc_csv = base.parent / f"{base.name}_PSC_Log.csv"
            write_dict_rows_csv(psc_log_rows, psc_csv, log_cb, progress_cb, written_offset, write_total, control)
            written_offset += len(psc_log_rows)
            csv_files.append(str(psc_csv))
        if psc_param_rows:
            psc_param_csv = base.parent / f"{base.name}_PSC_Params.csv"
            write_dict_rows_csv(psc_param_rows, psc_param_csv, log_cb, progress_cb, written_offset, write_total, control)
            written_offset += len(psc_param_rows)
            csv_files.append(str(psc_param_csv))
        if control:
            control.check()
        if progress_cb:
            progress_cb(write_total, write_total, "Creating summary workbook...")
        build_summary_only_xlsx(sorted_records, csa_hits, result, options, output_path, csv_files, log_cb)
        return

    outlog("Creating workbook...")
    # Approximate progress units for xlsx creation.  The final wb.save() step is
    # a ZIP compression phase that openpyxl cannot subdivide, so reserve a small
    # tail range for it and report row-by-row progress before that phase.
    est_error_count = sum(1 for r in records if r.category in {"ERROR", "WARNING"} or r.level in {"ERROR", "ERR", "WARN", "WARNING", "FATAL", "ALARM"})
    est_type_rows = 0 if len(records) > 200000 else len(records)
    est_merged_rows = 1 if len(records) > 100000 else len(records)
    write_total = max(1, len(records) + est_error_count + len(csa_hits) + est_type_rows + est_merged_rows + len(psc_log_rows or []) + len(psc_param_rows or []) + 5000)
    write_cursor = 0
    fast_xlsx_mode = bool(getattr(options, "fast_xlsx_mode", True))
    timestamp_as_text = bool(getattr(options, "timestamp_as_text", True))
    disable_styles_large = bool(getattr(options, "disable_styles_large", True))
    no_styles = bool(disable_styles_large and len(records) >= 50000)

    def write_progress(label: str):
        if progress_cb:
            progress_cb(min(write_cursor, write_total), write_total, label)

    def write_sheet_rows(ws, headers, rows, label: str):
        nonlocal write_cursor
        n = write_rows(ws, headers, rows, progress_cb=progress_cb, progress_start=write_cursor, progress_total=write_total, progress_label=label, control=control, styles=not no_styles, timestamp_format=not timestamp_as_text)
        write_cursor += max(1, n)
        write_progress(f"{label}: completed")
        return n

    def write_generic_progress(title: str, headers, rows, table_name: str):
        nonlocal write_cursor
        wsx = wb.create_sheet(title)
        n = write_rows(wsx, headers, rows, progress_cb=progress_cb, progress_start=write_cursor, progress_total=write_total, progress_label=f"Writing {title}", control=control, styles=not no_styles, timestamp_format=not timestamp_as_text)
        write_cursor += max(1, n)
        apply_table(wsx, table_name)
        autosize(wsx)
        write_progress(f"Writing {title}: completed")
        return wsx

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"] = "Log Merge Tool - No Excel"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")

    summary_rows = [
        ["Version", APP_VERSION],
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Source Folder", options.source_folder],
        ["Output Folder", options.output_folder],
        ["Recursive", "Yes" if options.recursive else "No"],
        ["Start Date", options.start_date if options.use_start else "Not specified"],
        ["End Date", options.end_date if options.use_end else "Not specified"],
        ["Serial Number", options.serial],
        ["Site Name", options.site],
        ["Scanned Files", result.scanned_files],
        ["Parsed Files", result.parsed_files],
        ["Skipped Files", result.skipped_files],
        ["Total Records", result.total_records],
        ["Included Records", result.included_records],
        ["CSA Error Count", result.csa_error_count],
    ]
    for r, row in enumerate(summary_rows, start=3):
        ws.cell(r, 1, row[0]).font = Font(bold=True)
        ws.cell(r, 2, row[1])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80

    # Search sheet: all merged records, sorted descending by timestamp.
    outlog(f"Sorting and writing merged records: {len(records)} rows")
    sorted_records = sorted(records, key=lambda r: r.timestamp or datetime.min, reverse=True)
    headers = record_output_headers()
    rows = [record_to_output_row_fast(r, timestamp_as_text=timestamp_as_text) for r in sorted_records]
    ws_search = wb.create_sheet("Search")
    outlog("Writing Search sheet rows...")
    write_sheet_rows(ws_search, headers, rows, "Writing Search sheet")
    outlog("Applying Search sheet timestamp format, C2 freeze, filter, and width settings...")
    apply_table(ws_search, "SearchTable")
    if not fast_xlsx_mode:
        autosize(ws_search)
    ws_search.column_dimensions["A"].width = 24
    ws_search.column_dimensions["G"].width = 80
    ws_search.column_dimensions["H"].width = 100
    activate_search_sheet(wb, ws_search)

    outlog("Writing PSC detail sheets...")
    # PSC sheets: merge-capable PSC log import, equivalent to ReadPSClog_LogMergeReady output.
    psc_log_rows = psc_log_rows or []
    psc_param_rows = psc_param_rows or []
    if psc_log_rows:
        max_param = max((int(r.get("ParamCount", 0)) for r in psc_log_rows), default=0)
        psc_headers = ["SourceFile", "LogDate", "Time", "DateTime", "Process", "Module", "Function", "Message", "ParameterText", "ParamCount", "RawLine"] + [f"Param{i}" for i in range(1, max_param + 1)]
        write_generic_progress("PSC_Log", psc_headers, [[r.get(h, "") for h in psc_headers] for r in psc_log_rows], "PSCLogTable")
    if psc_param_rows:
        psc_param_headers = ["SourceFile", "LogDate", "Time", "DateTime", "Process", "Module", "Function", "ParameterName", "ParameterValue", "Message", "RawLine"]
        write_generic_progress("PSC_Params", psc_param_headers, [[r.get(h, "") for h in psc_param_headers] for r in psc_param_rows], "PSCParamTable")

    outlog("Writing Errors sheet...")
    # Errors sheet.
    error_records = [r for r in sorted_records if r.category in {"ERROR", "WARNING"} or r.level in {"ERROR", "ERR", "WARN", "WARNING", "FATAL", "ALARM"}]
    ws_errors = wb.create_sheet("Errors")
    write_sheet_rows(ws_errors, headers, [record_to_output_row_fast(r, timestamp_as_text=timestamp_as_text) for r in error_records], "Writing Errors sheet")
    apply_table(ws_errors, "ErrorsTable")
    autosize(ws_errors)
    if ws_errors.max_row >= 2 and not no_styles:
        ws_errors.conditional_formatting.add(f"A2:H{ws_errors.max_row}", FormulaRule(formula=['OR($E2="ERROR",$E2="ERR",$E2="FATAL",$E2="ALARM",$F2="ERROR")'], fill=PatternFill("solid", fgColor="F4CCCC")))

    # CSA Error List hit sheet.
    ws_csa_hits = wb.create_sheet("CSA Error Hits")
    csa_headers = ["Timestamp", "File", "Line", "Keyword", "Message", "Raw"]
    csa_rows = [csa_hit_to_output_row(h) for h in csa_hits]
    write_rows(ws_csa_hits, csa_headers, csa_rows)
    apply_table(ws_csa_hits, "CSAErrorHits")
    autosize(ws_csa_hits)

    outlog("Writing log-type sheets...")
    # Type sheets. Search always contains all records. For very large outputs,
    # full duplicate per-type sheets can make xlsx creation appear stuck.
    if len(sorted_records) > 200000 or fast_xlsx_mode:
        ws_type_summary = wb.create_sheet("Type_Summary")
        summary = []
        for typ in ["WS", "CGA", "CSA", "PSC", "MRSERVER", "GESYS", "LAIS", "UNKNOWN"]:
            cnt = sum(1 for r in sorted_records if r.source_type == typ)
            if cnt:
                summary.append([typ, cnt, "Full records are in Search sheet"])
        write_sheet_rows(ws_type_summary, ["SourceType", "RecordCount", "Note"], summary, "Writing Type Summary")
        apply_table(ws_type_summary, "TypeSummaryTable")
        autosize(ws_type_summary)
    else:
        for typ in ["WS", "CGA", "CSA", "PSC", "MRSERVER", "GESYS", "LAIS", "UNKNOWN"]:
            typ_records = [r for r in sorted_records if r.source_type == typ]
            if not typ_records:
                continue
            ws_typ = wb.create_sheet(typ)
            write_sheet_rows(ws_typ, headers, [record_to_output_row(r) for r in typ_records], f"Writing {typ} sheet")
            apply_table(ws_typ, f"{typ}Table")
            autosize(ws_typ)

    outlog("Writing Merged_1 compatibility sheet...")
    # Merged_1 compatibility sheet. For very large outputs, duplicating the full
    # Search sheet can double the save time and file size. Keep a lightweight
    # compatibility sheet instead; Search contains the complete merged data.
    ws_merged = wb.create_sheet("Merged_1")
    if len(rows) > 100000 or bool(getattr(options, "skip_merged1", True)):
        write_sheet_rows(ws_merged, ["Notice"], [["Full merged records are in the Search sheet. Merged_1 full duplicate was skipped for performance because the output exceeded 100,000 rows."]], "Writing Merged_1 notice")
    else:
        write_sheet_rows(ws_merged, headers, rows, "Writing Merged_1 compatibility sheet")
        apply_table(ws_merged, "MergedTable")
        autosize(ws_merged)

    # Config sheets for traceability.
    ws_rules = wb.create_sheet("CSA Error List")
    write_rows(ws_rules, ["Keyword", "Message"], [[r.get("keyword", ""), r.get("message", "")] for r in load_json_file("csa_error_rules.json", DEFAULT_CSA_RULES)])
    apply_table(ws_rules, "CSAErrorList")
    autosize(ws_rules)

    ws_lists = wb.create_sheet("Lists")
    write_rows(ws_lists, ["Serial Number", "Site Name"], [[r.get("serial", ""), r.get("site", "")] for r in load_json_file("site_serial_map.json", DEFAULT_SITE_MAP)])
    apply_table(ws_lists, "SiteSerialList")
    autosize(ws_lists)

    # Ensure the user opens directly on Search with C2 freeze/filter applied.
    try:
        activate_search_sheet(wb, ws_search)
    except Exception:
        pass

    outlog("Saving xlsx file... Please wait.")
    if progress_cb:
        progress_cb(max(0, write_total - 5000), write_total, "Finalizing workbook and compressing xlsx... Excel library cannot report sub-progress in this step.")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    if progress_cb:
        progress_cb(write_total - 1000, write_total, "Validating xlsx package...")
    strip_excel_table_xml(output_path)
    if progress_cb:
        progress_cb(write_total, write_total, "xlsx save completed.")
    outlog("xlsx save completed.")



def parse_candidate_file(args) -> tuple[Path, Optional[str], int, list[LogRecord], list[dict[str, Any]], list[dict[str, Any]], Optional[str]]:
    p, include_unknown = args
    typ = classify_file(p, include_unknown)
    if typ is None:
        return p, None, 0, [], [], [], None
    try:
        if typ == "PSC":
            raw_psc_log, raw_psc_params, parsed = parse_psc_file_detail(p)
        else:
            raw_psc_log, raw_psc_params = [], []
            parsed = parse_file(p, typ)
        return p, typ, len(parsed), parsed, raw_psc_log, raw_psc_params, None
    except Exception as exc:
        return p, typ, 0, [], [], [], str(exc)

def run_merge(options: RunOptions, progress_cb=None, log_cb=None, control=None) -> RunResult:
    def log(msg: str):
        if log_cb:
            log_cb(msg)

    src = Path(options.source_folder)
    out_dir = Path(options.output_folder) if options.output_folder else src
    if not src.exists():
        raise FileNotFoundError(f"Source folder not found: {src}")
    if not out_dir.exists():
        out_dir.mkdir(parents=True, exist_ok=True)

    start = date.fromisoformat(options.start_date) if options.use_start and options.start_date else None
    end = date.fromisoformat(options.end_date) if options.use_end and options.end_date else None
    if start and end and start > end:
        raise ValueError("Start Date is later than End Date.")

    enabled = enabled_types(options)
    if not enabled:
        raise ValueError("No log type is selected.")

    if getattr(options, "selected_files", None):
        files = [Path(x) for x in options.selected_files if Path(x).exists()]
        log(f"Using Smart File Discovery selected files: {len(files)}")
    else:
        files = list(iter_files(src, options.recursive))
    result = RunResult(scanned_files=len(files))
    records: list[LogRecord] = []
    psc_log_rows: list[dict[str, Any]] = []
    psc_param_rows: list[dict[str, Any]] = []
    source_entries: list[dict[str, Any]] = []

    log(f"Scanned files: {len(files)}")
    log("v15 mode: Date filter uses content timestamp first, filename date fallback only, never file created/modified date. XLSX writer removes all Excel Table XML and uses AutoFilter only.")
    # v6 Turbo parser: parse files concurrently. This is mostly I/O-bound, so
    # ThreadPoolExecutor is effective and avoids multiprocessing freeze issues in onefile EXEs.
    max_workers = int(getattr(options, "worker_count", 0) or 0)
    if max_workers <= 0:
        try:
            max_workers = min(8, max(2, (os.cpu_count() or 4)))
        except Exception:
            max_workers = 4
    log(f"Turbo parser workers: {max_workers}")
    completed = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {executor.submit(parse_candidate_file, (p, options.include_unknown)): p for p in files}
        for fut in concurrent.futures.as_completed(future_map):
            if control:
                control.check()
            p = future_map[fut]
            completed += 1
            if progress_cb:
                progress_cb(completed, len(files), p.name)
            try:
                p, typ, parsed_count, parsed, raw_psc_log, raw_psc_params, err = fut.result()
            except Exception as exc:
                result.skipped_files += 1
                log(f"Skipped unreadable file: {p.name} / {exc}")
                continue
            if typ is None or typ not in enabled:
                result.skipped_files += 1
                continue
            if err:
                result.skipped_files += 1
                log(f"Skipped unreadable file: {p.name} / {err}")
                continue
            result.parsed_files += 1
            before_total = result.total_records
            before_included = result.included_records
            file_first_ts = min((r.timestamp for r in parsed if r.timestamp), default=None)
            file_last_ts = max((r.timestamp for r in parsed if r.timestamp), default=None)
            for rec in parsed:
                result.total_records += 1
                if in_date_range(rec.timestamp, None, start, end):
                    records.append(rec)
                    result.included_records += 1
            try:
                stat = p.stat()
                file_size = stat.st_size
            except Exception:
                file_size = 0
            source_entries.append({
                "type": typ,
                "file_name": p.name,
                "path": str(p),
                "size_bytes": file_size,
                "sha256": safe_sha256(p),
                "first_timestamp": format_timestamp(file_first_ts),
                "last_timestamp": format_timestamp(file_last_ts),
                "rows_read": len(parsed),
                "rows_output": result.included_records - before_included,
                "noise_excluded_rows": 0,
            })
            if typ == "PSC":
                for r in raw_psc_log:
                    ts = parse_datetime_from_text(str(r.get("DateTime", "")))
                    if in_date_range(ts, None, start, end):
                        psc_log_rows.append(r)
                for r in raw_psc_params:
                    ts = parse_datetime_from_text(str(r.get("DateTime", "")))
                    if in_date_range(ts, None, start, end):
                        psc_param_rows.append(r)

    before_noise_count = len(records)
    records = apply_noise_to_records(records, options.noise_enabled_merge, options.noise_exclude_output, options.noise_learning, log)
    noise_removed_total = max(0, before_noise_count - len(records))
    if noise_removed_total:
        log(f"Noise excluded rows for manifest: {noise_removed_total}")
    rules = load_json_file("csa_error_rules.json", DEFAULT_CSA_RULES)
    csa_hits = check_csa_errors(records, rules)
    result.csa_error_count = len(csa_hits)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    site_part = re.sub(r"[^A-Za-z0-9_-]+", "_", options.site.strip()) if options.site.strip() else "NoSite"
    serial_part = re.sub(r"[^A-Za-z0-9_-]+", "_", options.serial.strip()) if options.serial.strip() else "NoSerial"
    output_path = out_dir / f"LogMerge_NoExcel_{site_part}_{serial_part}_{stamp}.xlsx"
    log(f"Scan/parse completed. Parsed files: {result.parsed_files}, included records: {result.included_records}")
    if progress_cb:
        write_total = max(1, result.included_records + len(psc_log_rows) + len(psc_param_rows))
        progress_cb(0, write_total, "Writing output files...")
    log("Now writing output files. For large folders this step can take several minutes.")
    build_output(records, csa_hits, result, options, output_path, psc_log_rows, psc_param_rows, log, progress_cb, control)
    result.output_path = str(output_path)
    try:
        # Store final output row counts after Noise filtering.
        output_counts: dict[tuple[str, str], int] = {}
        for r in records:
            output_counts[(r.source_type, r.filename)] = output_counts.get((r.source_type, r.filename), 0) + 1
        for ent in source_entries:
            ent["rows_output_after_noise"] = output_counts.get((ent.get("type", ""), ent.get("file_name", "")), 0)
        manifest_path = write_merge_manifest(output_path, options, result, source_entries, [], log)
        result.output_path = str(output_path)
        log(f"Merge manifest: {manifest_path}")
    except Exception:
        log("Manifest write failed:\n" + traceback.format_exc())
    return result


class RunControl:
    def __init__(self):
        self._cancel = threading.Event()
        self._pause = threading.Event()

    def request_cancel(self):
        self._cancel.set()

    def set_paused(self, paused: bool):
        if paused:
            self._pause.set()
        else:
            self._pause.clear()

    def check(self):
        if self._cancel.is_set():
            raise RuntimeError("Operation cancelled by user.")
        while self._pause.is_set():
            if self._cancel.is_set():
                raise RuntimeError("Operation cancelled by user.")
            time.sleep(0.15)


class MergeWorker(QThread):
    progress = Signal(int, int, str)
    log = Signal(str)
    finished_ok = Signal(object)
    failed = Signal(str)

    def __init__(self, options: RunOptions):
        super().__init__()
        self.options = options
        self.control = RunControl()

    def request_cancel(self):
        self.control.request_cancel()

    def set_paused(self, paused: bool):
        self.control.set_paused(paused)

    def run(self):
        try:
            result = run_merge(self.options, self.progress.emit, self.log.emit, self.control)
            self.finished_ok.emit(result)
        except Exception as exc:
            if "cancelled by user" in str(exc).lower():
                self.failed.emit("Operation cancelled by user.")
            else:
                self.failed.emit(traceback.format_exc())



class SimpleTaskWorker(QThread):
    finished_ok = Signal(object)
    failed = Signal(str)

    def __init__(self, func):
        super().__init__()
        self.func = func

    def run(self):
        try:
            self.finished_ok.emit(self.func())
        except Exception:
            self.failed.emit(traceback.format_exc())



# ---------------------------------------------------------------------------
# v17 Smart Noise Learning
# ---------------------------------------------------------------------------
NOISE_DB_NAME = "NoiseDB.sqlite"
NOISE_SUGGEST_LIMIT = 200
NOISE_KEYWORDS = [
    "heartbeat", "heart beat", "keepalive", "keep alive", "poll", "polling",
    "status ok", "alive", "timer", "tick", "debug trace", "periodic",
    "no change", "idle", "socket select", "refresh", "update display",
]


def runtime_data_dir() -> Path:
    base = os.environ.get("LOCALAPPDATA") or tempfile.gettempdir()
    d = Path(base) / "LogMergeTool_NoExcel"
    d.mkdir(parents=True, exist_ok=True)
    return d


def noise_db_path() -> Path:
    return runtime_data_dir() / NOISE_DB_NAME


def init_noise_db() -> None:
    con = sqlite3.connect(noise_db_path())
    try:
        cur = con.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS noise_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pattern TEXT NOT NULL,
                match_type TEXT NOT NULL DEFAULT 'contains',
                log_type TEXT NOT NULL DEFAULT 'ANY',
                action TEXT NOT NULL DEFAULT 'hide',
                scope TEXT NOT NULL DEFAULT 'both',
                enabled INTEGER NOT NULL DEFAULT 1,
                approved INTEGER NOT NULL DEFAULT 1,
                confidence REAL NOT NULL DEFAULT 1.0,
                hit_count INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                notes TEXT DEFAULT ''
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS noise_samples (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                log_type TEXT NOT NULL,
                message TEXT NOT NULL,
                raw TEXT NOT NULL,
                label TEXT NOT NULL,
                score REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            )
            """
        )
        con.commit()
    finally:
        con.close()


def _compact_pattern(text: str) -> str:
    t = re.sub(r"\s+", " ", text or "").strip()
    # Remove highly variable values to make suggested rules reusable but safe.
    t = re.sub(r"\b20\d{2}[-/]\d{1,2}[-/]\d{1,2}\b", "", t)
    t = re.sub(r"\b\d{1,2}:\d{2}:\d{2}(?:[.:]\d+)?\b", "", t)
    t = re.sub(r"\b0x[0-9a-fA-F]+\b", "", t)
    t = re.sub(r"\b\d{5,}\b", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    if len(t) > 120:
        t = t[:120].rstrip()
    return t or (text or "")[:80]



def normalize_noise_text(text: str) -> str:
    """Normalize log text for operator Noise Rule matching.

    Real logs often contain tabs, doubled spaces, control characters, or values
    split differently between Message and Raw.  Operator rules should match the
    visible text even when whitespace differs.
    """
    if text is None:
        return ""
    t = str(text)
    t = clean_excel_value(t)
    t = t.replace("\t", " ").replace("\r", " ").replace("\n", " ")
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t


def noise_record_text(record: LogRecord) -> tuple[str, str]:
    """Return raw and normalized searchable text for Noise matching."""
    parts = [
        record.message or "",
        record.raw or "",
        viewer_message_body(record.message or ""),
        viewer_message_body(record.raw or ""),
        record.level or "",
        record.category or "",
    ]
    raw_text = "\n".join(str(x) for x in parts if x is not None)
    return raw_text, normalize_noise_text(raw_text)


def increment_noise_hits(rule_ids: Iterable[int]) -> None:
    ids = [int(x) for x in rule_ids if x is not None]
    if not ids:
        return
    con = sqlite3.connect(noise_db_path())
    try:
        now = datetime.now().isoformat(timespec="seconds")
        for rid in ids:
            con.execute("UPDATE noise_rules SET hit_count = COALESCE(hit_count,0) + 1, updated_at=? WHERE id=?", (now, rid))
        con.commit()
    finally:
        con.close()

class NoiseEngine:
    def __init__(self):
        init_noise_db()
        self.rules = self.load_rules()

    def load_rules(self) -> list[dict[str, Any]]:
        con = sqlite3.connect(noise_db_path())
        con.row_factory = sqlite3.Row
        try:
            rows = con.execute(
                "SELECT * FROM noise_rules WHERE enabled=1 AND approved=1 ORDER BY id"
            ).fetchall()
            return [dict(r) for r in rows]
        finally:
            con.close()

    def reload(self):
        self.rules = self.load_rules()

    def match_rule(self, record: LogRecord) -> Optional[dict[str, Any]]:
        raw_text, text_n = noise_record_text(record)
        typ = (record.source_type or "").upper()
        for rule in self.rules:
            rt = str(rule.get("log_type") or "ANY").upper()
            if rt not in ("ANY", typ):
                continue
            pat = str(rule.get("pattern") or "")
            if not pat:
                continue
            pat_n = normalize_noise_text(pat)
            mt = str(rule.get("match_type") or "contains").lower()
            try:
                if mt == "regex":
                    # Regex is evaluated against both original and normalized text.
                    if re.search(pat, raw_text, re.I) or (pat_n and re.search(re.escape(pat_n), text_n, re.I)):
                        return rule
                elif mt == "startswith":
                    if text_n.startswith(pat_n):
                        return rule
                elif mt == "endswith":
                    if text_n.endswith(pat_n):
                        return rule
                else:
                    # Operator expectation: visible text should match even when
                    # whitespace/control characters differ.  Use normalized contains.
                    if pat_n and pat_n in text_n:
                        return rule
            except re.error:
                continue
        return None

    def is_noise(self, record: LogRecord) -> bool:
        return self.match_rule(record) is not None

    def score(self, record: LogRecord) -> int:
        text = ((record.message or "") + " " + (record.raw or "")).lower()
        score = 0
        if any(k in text for k in NOISE_KEYWORDS):
            score += 55
        if (record.level or "").upper() in {"DBG", "DEBUG", "TRACE"}:
            score += 15
        if len((record.message or record.raw or "").strip()) < 35:
            score += 10
        if re.search(r"\b(ok|normal|idle|ready|alive)\b", text):
            score += 10
        if re.search(r"\b(error|fail|failed|warn|alarm|exception|abort|timeout)\b", text):
            score -= 45
        return max(0, min(100, score))

    def add_sample(self, record: LogRecord, label: str, score: int = 0) -> None:
        con = sqlite3.connect(noise_db_path())
        try:
            con.execute(
                "INSERT INTO noise_samples(log_type,message,raw,label,score,created_at) VALUES(?,?,?,?,?,?)",
                (record.source_type, record.message, record.raw, label, float(score), datetime.now().isoformat(timespec="seconds")),
            )
            con.commit()
        finally:
            con.close()

    def add_rule(self, pattern: str, log_type: str = "ANY", match_type: str = "contains", action: str = "hide", scope: str = "both", notes: str = "Approved by operator") -> int:
        pattern = re.sub(r"\s+", " ", str(pattern or "")).strip()
        now = datetime.now().isoformat(timespec="seconds")
        con = sqlite3.connect(noise_db_path())
        try:
            cur = con.cursor()
            cur.execute(
                """
                INSERT INTO noise_rules(pattern,match_type,log_type,action,scope,enabled,approved,confidence,created_at,updated_at,notes)
                VALUES(?,?,?,?,?,1,1,1.0,?,?,?)
                """,
                (pattern, match_type, (log_type or "ANY").upper(), action, scope, now, now, notes),
            )
            con.commit()
            rid = int(cur.lastrowid)
        finally:
            con.close()
        self.reload()
        return rid

    def add_rule_from_record(self, record: LogRecord, match_type: str = "contains", action: str = "hide", scope: str = "both") -> int:
        pattern = _compact_pattern(record.message or record.raw)
        rid = self.add_rule(pattern, (record.source_type or "ANY").upper(), match_type, action, scope, "Approved by operator from selected row")
        self.add_sample(record, "noise-approved", 100)
        return rid

    def mark_not_noise(self, record: LogRecord) -> None:
        self.add_sample(record, "not-noise", 0)


def apply_noise_to_records(records: list[LogRecord], enabled: bool, exclude_output: bool, learning: bool, log_cb=None) -> list[LogRecord]:
    """Apply Smart Noise rules.

    v25 behavior:
    - If enabled=True, approved/enabled rules are applied to the merge result and matching rows are removed.
      This matches the operator expectation that "Enable in Merge" actually filters the output.
    - exclude_output is kept for compatibility with older projects; either flag removes matching rows.
    - Learning mode only saves suggestions and never removes records by itself.
    """
    if not (enabled or exclude_output or learning):
        return records
    engine = NoiseEngine()
    kept: list[LogRecord] = []
    removed = 0
    suggested = 0
    do_remove = bool(enabled or exclude_output)
    matched_rule_ids: list[int] = []
    for r in records:
        matched_rule = engine.match_rule(r) if do_remove else None
        if matched_rule is not None:
            rid = matched_rule.get("id")
            if rid is not None:
                matched_rule_ids.append(int(rid))
            removed += 1
            continue
        kept.append(r)
        if learning and suggested < NOISE_SUGGEST_LIMIT and not engine.is_noise(r):
            sc = engine.score(r)
            if sc >= 80:
                engine.add_sample(r, "suggest", sc)
                suggested += 1
    increment_noise_hits(matched_rule_ids)
    if log_cb:
        if removed:
            log_cb(f"Noise Filter: excluded {removed} approved noise rows from merge output.")
        elif do_remove:
            log_cb("Noise Filter: enabled, but no rows matched approved enabled rules.")
        if suggested:
            log_cb(f"Noise Learning: saved {suggested} candidate rows for operator review. No candidate was removed automatically.")
    return kept

class NoiseRuleManagerDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        init_noise_db()
        self.setWindowTitle("Noise Rules Manager")
        self.resize(980, 520)
        root = QVBoxLayout(self)
        note = QLabel("Approved rules are operator-controlled. Disable or delete a rule to restore matching logs.")
        root.addWidget(note)
        self.table = QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels(["Status", "ID", "Enabled", "LogType", "Match", "Pattern", "Action", "Scope", "Hits", "Updated"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.currentCellChanged.connect(lambda *_: self.update_toggle_button_text())
        root.addWidget(self.table)
        buttons = QHBoxLayout()
        self.refresh_btn = QPushButton("Refresh")
        self.add_btn = QPushButton("Add Manual Rule")
        self.toggle_btn = QPushButton("Enable/Disable")
        self.delete_btn = QPushButton("Delete")
        self.close_btn = QPushButton("Close")
        for b in [self.refresh_btn, self.add_btn, self.toggle_btn, self.delete_btn]:
            buttons.addWidget(b)
        buttons.addStretch(1); buttons.addWidget(self.close_btn)
        root.addLayout(buttons)
        self.refresh_btn.clicked.connect(self.load_rules)
        self.add_btn.clicked.connect(self.add_manual_rule)
        self.toggle_btn.clicked.connect(self.toggle_selected)
        self.delete_btn.clicked.connect(self.delete_selected)
        self.close_btn.clicked.connect(self.accept)
        self.load_rules()

    def selected_rule_id(self) -> Optional[int]:
        row = self.table.currentRow()
        if row < 0:
            return None
        item = self.table.item(row, 1)
        if not item:
            return None
        try:
            return int(item.text())
        except Exception:
            return None

    def load_rules(self):
        con = sqlite3.connect(noise_db_path())
        con.row_factory = sqlite3.Row
        try:
            rows = con.execute("SELECT * FROM noise_rules ORDER BY id DESC LIMIT 1000").fetchall()
        finally:
            con.close()
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            enabled_text = "Yes" if row["enabled"] else "No"
            status_text = "Enabled" if row["enabled"] else "Disabled"
            values = [
                status_text, row["id"], enabled_text, row["log_type"], row["match_type"], row["pattern"],
                row["action"], row["scope"], row["hit_count"], row["updated_at"],
            ]
            for c, val in enumerate(values):
                self.table.setItem(r, c, QTableWidgetItem(str(val)))
        self.table.resizeColumnsToContents()
        self.update_toggle_button_text()

    def update_toggle_button_text(self):
        row = self.table.currentRow()
        if row < 0:
            self.toggle_btn.setText("Enable/Disable Rule")
            return
        enabled_item = self.table.item(row, 2)
        enabled = enabled_item and enabled_item.text().lower() == "yes"
        self.toggle_btn.setText("Disable Rule" if enabled else "Enable Rule")

    def add_manual_rule(self):
        pattern, ok = QInputDialog.getText(self, "Add Manual Noise Rule", "Text to match (contains):")
        if not ok or not pattern.strip():
            return
        log_type, ok = QInputDialog.getText(self, "Log Type", "Log type (ANY/WS/CGA/CSA/MRSERVER/GESYS/LAIS/PSC/REVIEW):", text="ANY")
        if not ok:
            return
        now = datetime.now().isoformat(timespec="seconds")
        con = sqlite3.connect(noise_db_path())
        try:
            con.execute(
                "INSERT INTO noise_rules(pattern,match_type,log_type,action,scope,enabled,approved,confidence,created_at,updated_at,notes) VALUES(?,?,?,?,?,1,1,1.0,?,?,?)",
                (pattern.strip(), "contains", (log_type or "ANY").strip().upper(), "hide", "both", now, now, "Manual rule"),
            )
            con.commit()
        finally:
            con.close()
        self.load_rules()

    def toggle_selected(self):
        rid = self.selected_rule_id()
        if rid is None:
            return
        con = sqlite3.connect(noise_db_path())
        try:
            con.execute("UPDATE noise_rules SET enabled = CASE enabled WHEN 1 THEN 0 ELSE 1 END, updated_at=? WHERE id=?", (datetime.now().isoformat(timespec="seconds"), rid))
            con.commit()
        finally:
            con.close()
        self.load_rules()

    def delete_selected(self):
        rid = self.selected_rule_id()
        if rid is None:
            return
        if QMessageBox.question(self, "Delete Rule", f"Delete rule ID {rid}?") != QMessageBox.Yes:
            return
        con = sqlite3.connect(noise_db_path())
        try:
            con.execute("DELETE FROM noise_rules WHERE id=?", (rid,))
            con.commit()
        finally:
            con.close()
        self.load_rules()

# ---------------------------------------------------------------------------
# v16 Dual Log Viewer
# ---------------------------------------------------------------------------
VIEWER_COLUMNS = ["Timestamp", "Message", "SourceType", "File", "Line", "Level", "Category"]


def format_viewer_timestamp(ts: Optional[datetime]) -> str:
    if not ts:
        return ""
    # Display as explicit one-column timestamp. Excel-like millisecond/decisecond
    # values are kept compact but zero-padded.
    frac = int(round(ts.microsecond / 100000.0))
    if frac >= 10:
        ts = ts + timedelta(seconds=1)
        frac = 0
    return ts.strftime("%Y/%m/%d %H:%M:%S") + f".{frac}"




def parse_viewer_datetime_text(text: str, *, is_end: bool = False) -> Optional[datetime]:
    """Parse viewer time filter text.

    Supported examples:
      2026/07/04 13:07:58.6
      2026-07-04 13:07:58
      2026/07/04

    Date-only values become start-of-day for Start and end-of-day for End.
    """
    t = (text or "").strip()
    if not t:
        return None
    t = t.replace("T", " ").replace("-", "/")
    # Date only
    m = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})", t)
    if m:
        y, mo, d = map(int, m.groups())
        if is_end:
            return datetime(y, mo, d, 23, 59, 59, 999999)
        return datetime(y, mo, d, 0, 0, 0, 0)
    # Date + time. Accept fractional seconds of any length; pad/truncate to microseconds.
    m = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})\s+(\d{1,2}):(\d{1,2})(?::(\d{1,2})(?:[\.:](\d{1,6}))?)?", t)
    if not m:
        raise ValueError(f"Invalid date/time: {text}")
    y, mo, d, hh, mm, ss, frac = m.groups()
    us = 0
    if frac:
        us = int((frac + "000000")[:6])
    return datetime(int(y), int(mo), int(d), int(hh), int(mm), int(ss or 0), us)

def record_to_viewer_row(r: LogRecord) -> dict[str, Any]:
    row = {
        "Timestamp": format_viewer_timestamp(r.timestamp),
        "SourceType": r.source_type,
        "File": r.filename,
        "Line": r.line_no,
        "Level": r.level,
        "Category": r.category,
        "Message": viewer_message_body(r.message),
        "Raw": clean_excel_value(r.raw),
        "_ts": r.timestamp,
    }
    # Review rows keep all extracted parameters as real table columns in the
    # Explorer. This makes review.out / review.out.ar usable as a parameter
    # matrix while other log types keep the normal compact layout.
    if r.source_type == "REVIEW" and r.raw:
        try:
            data = json.loads(r.raw)
            if isinstance(data, dict):
                for k, v in data.items():
                    if k not in {"date", "time"}:
                        row[str(k)] = clean_excel_value(v)
        except Exception:
            pass
    return row


def review_rows_to_viewer_records(path: Path) -> list[LogRecord]:
    records: list[LogRecord] = []
    for idx, row in enumerate(parse_reviewout_file(path), start=1):
        text_dt = " ".join(str(row.get(k, "")) for k in ("date", "time") if row.get(k, ""))
        ts = parse_datetime_from_text(text_dt) if text_dt.strip() else None
        msg_parts = []
        for k, v in row.items():
            if k in {"date", "time"}:
                continue
            if v not in (None, ""):
                msg_parts.append(f"{k}={v}")
            if len(msg_parts) >= 4:
                break
        msg = "; ".join(msg_parts) if msg_parts else "Review scan row"
        records.append(LogRecord(ts, "REVIEW", path.name, idx, "", "Review", msg, json.dumps(row, ensure_ascii=False)))
    return records


class LogTableModel(QAbstractTableModel):
    def __init__(self, rows: Optional[list[dict[str, Any]]] = None):
        super().__init__()
        self.rows = rows or []
        self.columns = list(VIEWER_COLUMNS)

    def rowCount(self, parent=QModelIndex()):
        return len(self.rows)

    def columnCount(self, parent=QModelIndex()):
        return len(self.columns)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or index.row() >= len(self.rows):
            return None
        row = self.rows[index.row()]
        col = self.columns[index.column()]
        if role == Qt.DisplayRole:
            return str(row.get(col, ""))
        if role == Qt.ToolTipRole:
            return str(row.get("Raw", row.get("Message", "")))
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return self.columns[section]
        return str(section + 1)

    def set_rows(self, rows: list[dict[str, Any]], columns: Optional[list[str]] = None):
        self.beginResetModel()
        self.rows = rows
        if columns is not None:
            self.columns = columns
        else:
            extra: list[str] = []
            seen = set(VIEWER_COLUMNS) | {"Raw", "_ts"}
            for row in rows[:500]:
                for key in row.keys():
                    if key not in seen and key not in extra:
                        extra.append(key)
            self.columns = list(VIEWER_COLUMNS) + extra
        self.endResetModel()

    def row_at(self, row: int) -> Optional[dict[str, Any]]:
        if 0 <= row < len(self.rows):
            return self.rows[row]
        return None


class DualLogViewer(QWidget):
    """Lightweight two-pane log viewer with timestamp cross reference.

    The viewer uses QTableView + QAbstractTableModel, so it does not create
    thousands of cell widgets. This keeps scrolling reasonably light even with
    large log lists. Data is loaded per side only when requested.
    """
    SOURCES = ["Merged", "WS", "CGA", "CSA", "MRSERVER", "GESYS", "LAIS", "PSC", "Review", "Custom File"]

    def __init__(self, parent_window: 'MainWindow'):
        super().__init__()
        self.parent_window = parent_window
        self.setWindowTitle("Dual Log Viewer - Cross Reference")
        self.resize(1300, 780)
        self._syncing = False
        self.left_model = LogTableModel([])
        self.right_model = LogTableModel([])
        self.left_ts: list[tuple[datetime, int]] = []
        self.right_ts: list[tuple[datetime, int]] = []
        self.left_all_rows: list[dict[str, Any]] = []
        self.right_all_rows: list[dict[str, Any]] = []
        self.build_ui()

    def build_ui(self):
        root = QVBoxLayout(self)

        # Global viewer controls stay in the center/top, while left/right file
        # selectors are placed directly above each pane so the operation target
        # is immediately obvious to the operator.
        top = QHBoxLayout()
        self.tolerance_spin = QSpinBox(); self.tolerance_spin.setRange(0, 3600); self.tolerance_spin.setValue(10); self.tolerance_spin.setSuffix(" sec")
        self.load_both_btn = QPushButton("Load Both")
        self.export_pair_btn = QPushButton("Export Selected Pair")
        self.apply_noise_chk = QCheckBox("Apply Noise Filter")
        self.apply_noise_chk.setChecked(True)
        self.noise_apply_mode = QComboBox(); self.noise_apply_mode.addItems(["Apply immediately after adding rule", "Apply only when Apply button is pressed"])
        self.apply_rules_btn = QPushButton("Apply Rules Now")
        self.not_noise_btn = QPushButton("Mark Selected Not Noise")
        self.manage_noise_btn = QPushButton("Manage Rules")
        self.status = QLabel("Ready")
        top.addWidget(QLabel("Jump tolerance:")); top.addWidget(self.tolerance_spin)
        top.addWidget(self.load_both_btn)
        top.addWidget(self.export_pair_btn)
        top.addWidget(self.apply_noise_chk)
        top.addWidget(self.noise_apply_mode)
        top.addWidget(self.apply_rules_btn)
        top.addWidget(self.not_noise_btn)
        top.addWidget(self.manage_noise_btn)
        top.addStretch(1)
        root.addLayout(top)

        # Viewer-only time range filter. This is independent from the Merge
        # date range and only changes what is displayed/searched in Explorer.
        time_bar = QHBoxLayout()
        self.viewer_time_enable = QCheckBox("Enable Viewer Time Range")
        self.viewer_start_edit = QLineEdit(); self.viewer_start_edit.setPlaceholderText("Start: 2026/07/04 13:07:58.6")
        self.viewer_end_edit = QLineEdit(); self.viewer_end_edit.setPlaceholderText("End: 2026/07/04 13:30:00.0")
        self.apply_time_left_btn = QPushButton("Apply to Left")
        self.apply_time_right_btn = QPushButton("Apply to Right")
        self.apply_time_both_btn = QPushButton("Apply to Both")
        self.clear_time_btn = QPushButton("Clear Time Range")
        self.set_start_btn = QPushButton("Set Start from Selected")
        self.set_end_btn = QPushButton("Set End from Selected")
        time_bar.addWidget(self.viewer_time_enable)
        time_bar.addWidget(self.viewer_start_edit, 1)
        time_bar.addWidget(self.viewer_end_edit, 1)
        time_bar.addWidget(self.apply_time_left_btn)
        time_bar.addWidget(self.apply_time_right_btn)
        time_bar.addWidget(self.apply_time_both_btn)
        time_bar.addWidget(self.clear_time_btn)
        time_bar.addWidget(self.set_start_btn)
        time_bar.addWidget(self.set_end_btn)
        root.addLayout(time_bar)

        self.left_source = QComboBox(); self.left_source.addItems(self.SOURCES); self.left_source.setCurrentText("Merged")
        self.right_source = QComboBox(); self.right_source.addItems(self.SOURCES); self.right_source.setCurrentText("PSC")
        self.load_left_btn = QPushButton("Load Left")
        self.load_right_btn = QPushButton("Load Right")
        self.left_search = QLineEdit(); self.left_search.setPlaceholderText("Search left message/raw...")
        self.right_search = QLineEdit(); self.right_search.setPlaceholderText("Search right message/raw...")
        self.find_left_btn = QPushButton("Find")
        self.find_right_btn = QPushButton("Find")
        self.mark_left_noise_btn = QPushButton("Add Left to Noise Rule")
        self.mark_right_noise_btn = QPushButton("Add Right to Noise Rule")
        self.copy_left_rule_btn = QPushButton("Copy Left Rule Text")
        self.copy_right_rule_btn = QPushButton("Copy Right Rule Text")
        self.left_file_label = QLabel("File: not loaded")
        self.right_file_label = QLabel("File: not loaded")
        self.left_file_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.right_file_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        # Tooltips for operator guidance.
        self.left_source.setToolTip("Select the log type shown in the LEFT pane. Use Custom File to load a single file without merge.")
        self.right_source.setToolTip("Select the log type shown in the RIGHT pane. PSC/Review can be compared without running merge.")
        self.load_left_btn.setToolTip("Load only the selected LEFT-side log. A progress popup is shown while loading.")
        self.load_right_btn.setToolTip("Load only the selected RIGHT-side log. A progress popup is shown while loading.")
        self.left_search.setToolTip("Search only the LEFT pane. It searches Message and Raw text.")
        self.right_search.setToolTip("Search only the RIGHT pane. It searches Message and Raw text.")
        self.find_left_btn.setToolTip("Find next matching row in the LEFT pane.")
        self.find_right_btn.setToolTip("Find next matching row in the RIGHT pane.")
        self.noise_apply_mode.setToolTip("Choose when newly approved Noise rules are applied to the Explorer view.")
        self.apply_rules_btn.setToolTip("Reload both panes and apply the current approved Noise rules.")
        self.load_both_btn.setToolTip("Load both left and right logs. Use this after selecting the two sources to compare.")
        self.tolerance_spin.setToolTip("Maximum time difference used when jumping to the nearest row on the other side.")
        self.export_pair_btn.setToolTip("Export the currently selected left/right row pair to CSV.")
        self.apply_noise_chk.setToolTip("Hide logs matching approved Smart Noise rules in this viewer only.")
        self.mark_left_noise_btn.setToolTip("Create a Noise rule from the selected LEFT row. The pattern is editable before approval.")
        self.mark_right_noise_btn.setToolTip("Create a Noise rule from the selected RIGHT row. The pattern is editable before approval.")
        self.copy_left_rule_btn.setToolTip("Copy a compact rule candidate from the selected LEFT row to the clipboard.")
        self.copy_right_rule_btn.setToolTip("Copy a compact rule candidate from the selected RIGHT row to the clipboard.")
        self.not_noise_btn.setToolTip("Mark the selected row as Not Noise so it can be reviewed later.")
        self.manage_noise_btn.setToolTip("Open Smart Noise rule manager. Rules can be enabled, disabled, or deleted later.")
        self.viewer_time_enable.setToolTip("Enable a Viewer-only start/end timestamp filter. It does not change Merge output files.")
        self.viewer_start_edit.setToolTip("Start timestamp for Explorer display/search. Example: 2026/07/04 13:07:58.6")
        self.viewer_end_edit.setToolTip("End timestamp for Explorer display/search. Date-only input is treated as end-of-day for End.")
        self.apply_time_left_btn.setToolTip("Apply the current time range only to the LEFT pane.")
        self.apply_time_right_btn.setToolTip("Apply the current time range only to the RIGHT pane.")
        self.apply_time_both_btn.setToolTip("Apply the current time range to both panes.")
        self.clear_time_btn.setToolTip("Clear the Viewer time range and restore all loaded rows.")
        self.set_start_btn.setToolTip("Copy the selected row timestamp into the Start field.")
        self.set_end_btn.setToolTip("Copy the selected row timestamp into the End field.")

        split = QSplitter(Qt.Horizontal)
        self.main_splitter = split
        split.setHandleWidth(10)
        split.setOpaqueResize(True)
        split.setChildrenCollapsible(False)
        split.setToolTip("Drag this center divider left or right to resize both log panes. Message columns expand or shrink automatically.")
        left_pane = QWidget(); left_layout = QVBoxLayout(left_pane); left_layout.setContentsMargins(4, 4, 4, 4)
        right_pane = QWidget(); right_layout = QVBoxLayout(right_pane); right_layout.setContentsMargins(4, 4, 4, 4)

        left_title = QLabel("LEFT LOG")
        left_title.setStyleSheet("font-weight:bold; background:#EAF3FF; padding:4px;")
        right_title = QLabel("RIGHT LOG")
        right_title.setStyleSheet("font-weight:bold; background:#EAF8EA; padding:4px;")
        left_layout.addWidget(left_title)
        right_layout.addWidget(right_title)

        left_controls = QHBoxLayout()
        left_controls.addWidget(QLabel("Left Log:")); left_controls.addWidget(self.left_source, 1)
        left_controls.addWidget(self.load_left_btn)
        left_layout.addLayout(left_controls)
        left_search_layout = QHBoxLayout()
        left_search_layout.addWidget(QLabel("Search:")); left_search_layout.addWidget(self.left_search, 1); left_search_layout.addWidget(self.find_left_btn)
        left_layout.addLayout(left_search_layout)
        left_layout.addWidget(self.left_file_label)
        left_noise_controls = QHBoxLayout()
        left_noise_controls.addWidget(self.mark_left_noise_btn)
        left_noise_controls.addWidget(self.copy_left_rule_btn)
        left_noise_controls.addStretch(1)
        left_layout.addLayout(left_noise_controls)

        right_controls = QHBoxLayout()
        right_controls.addWidget(QLabel("Right Log:")); right_controls.addWidget(self.right_source, 1)
        right_controls.addWidget(self.load_right_btn)
        right_layout.addLayout(right_controls)
        right_search_layout = QHBoxLayout()
        right_search_layout.addWidget(QLabel("Search:")); right_search_layout.addWidget(self.right_search, 1); right_search_layout.addWidget(self.find_right_btn)
        right_layout.addLayout(right_search_layout)
        right_layout.addWidget(self.right_file_label)
        right_noise_controls = QHBoxLayout()
        right_noise_controls.addWidget(self.mark_right_noise_btn)
        right_noise_controls.addWidget(self.copy_right_rule_btn)
        right_noise_controls.addStretch(1)
        right_layout.addLayout(right_noise_controls)

        self.left_table = QTableView(); self.left_table.setModel(self.left_model)
        self.right_table = QTableView(); self.right_table.setModel(self.right_model)
        for tbl in (self.left_table, self.right_table):
            tbl.setSelectionBehavior(QTableView.SelectRows)
            tbl.setSelectionMode(QTableView.SingleSelection)
            tbl.setAlternatingRowColors(True)
            tbl.setSortingEnabled(False)
            tbl.verticalHeader().setDefaultSectionSize(20)
            tbl.verticalHeader().setVisible(False)
            tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            # Default view: Timestamp is a fixed reference at the far left,
            # while Message stretches to use the available pane width. Dragging
            # the center splitter therefore makes the message window wider/narrower.
            tbl.setColumnWidth(0, 170)   # Timestamp
            tbl.setColumnWidth(1, 620)   # Message
            tbl.setColumnWidth(2, 90)    # SourceType
            tbl.setColumnWidth(3, 200)   # File
            tbl.setColumnWidth(4, 60)    # Line
            tbl.setColumnWidth(5, 70)    # Level
            tbl.setColumnWidth(6, 100)   # Category
            tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
            tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        left_layout.addWidget(self.left_table, 1)
        right_layout.addWidget(self.right_table, 1)
        split.addWidget(left_pane)
        split.addWidget(right_pane)
        split.setSizes([650, 650])
        root.addWidget(split, 1)

        self.detail = QTextEdit(); self.detail.setReadOnly(True); self.detail.setMaximumHeight(150)
        root.addWidget(self.detail)
        root.addWidget(self.status)

        self.load_left_btn.clicked.connect(lambda: self.load_side("left"))
        self.load_right_btn.clicked.connect(lambda: self.load_side("right"))
        self.load_both_btn.clicked.connect(self.load_both)
        self.export_pair_btn.clicked.connect(self.export_selected_pair)
        self.apply_noise_chk.toggled.connect(lambda _=False: self.apply_noise_rules_now())
        self.apply_rules_btn.clicked.connect(self.apply_noise_rules_now)
        self.left_search.returnPressed.connect(lambda: self.find_next("left"))
        self.right_search.returnPressed.connect(lambda: self.find_next("right"))
        self.find_left_btn.clicked.connect(lambda: self.find_next("left"))
        self.find_right_btn.clicked.connect(lambda: self.find_next("right"))
        self.mark_left_noise_btn.clicked.connect(lambda: self.approve_selected_noise("left"))
        self.mark_right_noise_btn.clicked.connect(lambda: self.approve_selected_noise("right"))
        self.copy_left_rule_btn.clicked.connect(lambda: self.copy_rule_text("left"))
        self.copy_right_rule_btn.clicked.connect(lambda: self.copy_rule_text("right"))
        self.not_noise_btn.clicked.connect(self.mark_selected_not_noise)
        self.manage_noise_btn.clicked.connect(self.manage_noise_rules)
        self.apply_time_left_btn.clicked.connect(lambda: self.apply_view_filters("left"))
        self.apply_time_right_btn.clicked.connect(lambda: self.apply_view_filters("right"))
        self.apply_time_both_btn.clicked.connect(lambda: (self.apply_view_filters("left"), self.apply_view_filters("right")))
        self.clear_time_btn.clicked.connect(self.clear_viewer_time_range)
        self.set_start_btn.clicked.connect(lambda: self.set_time_from_selected("start"))
        self.set_end_btn.clicked.connect(lambda: self.set_time_from_selected("end"))
        self.left_table.selectionModel().selectionChanged.connect(lambda *_: self.row_selected("left"))
        self.right_table.selectionModel().selectionChanged.connect(lambda *_: self.row_selected("right"))

    def log(self, text: str):
        self.status.setText(text)
        try:
            self.parent_window.log("[Viewer] " + text)
        except Exception:
            pass

    def make_progress(self, title: str, text: str) -> QProgressDialog:
        dlg = QProgressDialog(text, "Cancel", 0, 0, self)
        dlg.setWindowTitle(title)
        dlg.setWindowModality(Qt.ApplicationModal)
        dlg.setMinimumDuration(0)
        dlg.setAutoClose(False)
        dlg.setAutoReset(False)
        dlg.setValue(0)
        return dlg

    def load_both(self):
        self.load_side("left")
        self.load_side("right")

    def current_viewer_time_range(self) -> tuple[Optional[datetime], Optional[datetime]]:
        if not self.viewer_time_enable.isChecked():
            return None, None
        start = parse_viewer_datetime_text(self.viewer_start_edit.text(), is_end=False) if self.viewer_start_edit.text().strip() else None
        end = parse_viewer_datetime_text(self.viewer_end_edit.text(), is_end=True) if self.viewer_end_edit.text().strip() else None
        if start and end and start > end:
            raise ValueError("Viewer Start time is later than End time.")
        return start, end

    def apply_view_filters(self, side: str):
        """Apply Viewer-only time range to already loaded rows.

        This keeps the original loaded rows in memory so the operator can clear
        the filter without reloading files from disk. Search operates on the
        currently displayed rows after this filter is applied.
        """
        try:
            start, end = self.current_viewer_time_range()
        except Exception as e:
            QMessageBox.warning(self, "Viewer Time Range", str(e))
            return
        base_rows = self.left_all_rows if side == "left" else self.right_all_rows
        if start or end:
            rows = []
            for row in base_rows:
                ts = row.get("_ts")
                if not isinstance(ts, datetime):
                    continue
                if start and ts < start:
                    continue
                if end and ts > end:
                    continue
                rows.append(row)
        else:
            rows = list(base_rows)
        ts_index = [(row["_ts"], i) for i, row in enumerate(rows) if isinstance(row.get("_ts"), datetime)]
        ts_index.sort(key=lambda x: x[0])
        if side == "left":
            self.left_model.set_rows(rows)
            self.left_ts = ts_index
        else:
            self.right_model.set_rows(rows)
            self.right_ts = ts_index
        if start or end:
            self.log(f"Viewer time filter {side}: {len(rows)}/{len(base_rows)} rows")
        else:
            self.log(f"Viewer time filter cleared {side}: {len(rows)} rows")

    def clear_viewer_time_range(self):
        self.viewer_time_enable.setChecked(False)
        self.viewer_start_edit.clear()
        self.viewer_end_edit.clear()
        self.apply_view_filters("left")
        self.apply_view_filters("right")

    def set_time_from_selected(self, target: str):
        # Prefer the row with current focus; fall back to left, then right.
        for side, table, model in (("left", self.left_table, self.left_model), ("right", self.right_table, self.right_model)):
            if table.hasFocus() or table.selectionModel().selectedRows():
                indexes = table.selectionModel().selectedRows()
                if indexes:
                    row = model.row_at(indexes[0].row()) or {}
                    ts_text = row.get("Timestamp", "")
                    if ts_text:
                        if target == "start":
                            self.viewer_start_edit.setText(str(ts_text))
                        else:
                            self.viewer_end_edit.setText(str(ts_text))
                        self.viewer_time_enable.setChecked(True)
                        self.log(f"Set Viewer {target} from selected {side}: {ts_text}")
                        return
        QMessageBox.information(self, "Viewer Time Range", "Select a row first.")

    def source_to_records(self, source_name: str, progress: Optional[QProgressDialog] = None) -> list[LogRecord]:
        source_folder = Path(self.parent_window.source_edit.text().strip() or ".")
        recursive = self.parent_window.recursive_chk.isChecked()
        opt = self.parent_window.collect_options()
        if source_name == "Custom File":
            path_str, _ = QFileDialog.getOpenFileName(self, "Select log file", str(source_folder), "Log files (*.log *.txt *.out *.*);;All Files (*.*)")
            if not path_str:
                return []
            path = Path(path_str)
            if progress:
                progress.setLabelText(f"Loading custom file:\n{path.name}")
                QApplication.processEvents()
            if is_review_file(path.name):
                return review_rows_to_viewer_records(path)
            typ = classify_file(path, True) or "UNKNOWN"
            if typ == "PSC":
                _, _, recs = parse_psc_file_detail(path)
                return recs
            return parse_file(path, typ)

        if not source_folder.exists():
            QMessageBox.warning(self, "Dual Log Viewer", "Source Log Folder does not exist.")
            return []

        targets: set[str]
        if source_name == "Merged":
            targets = enabled_types(opt)
        elif source_name == "Review":
            targets = {"REVIEW"}
        else:
            targets = {source_name}

        # Build a candidate file list first so the progress popup can show a
        # useful percentage. This also avoids the previous impression that the
        # viewer had frozen while it was scanning a large folder.
        candidates: list[tuple[Path, str]] = []
        if progress:
            progress.setLabelText(f"Scanning files for {source_name} ...")
            QApplication.processEvents()
        for path in iter_files(source_folder, recursive):
            if progress and progress.wasCanceled():
                return []
            if source_name == "Review":
                if is_review_file(path.name):
                    candidates.append((path, "REVIEW"))
                continue
            typ = classify_file(path, opt.include_unknown)
            if typ and typ in targets:
                candidates.append((path, typ))

        if progress:
            progress.setRange(0, max(1, len(candidates)))
            progress.setValue(0)
            progress.setLabelText(f"Loading {source_name}: 0/{len(candidates)} files")
            QApplication.processEvents()

        records: list[LogRecord] = []
        for idx, (path, typ) in enumerate(candidates, start=1):
            if progress and progress.wasCanceled():
                self.log(f"Loading {source_name} canceled by operator.")
                break
            if progress:
                progress.setValue(idx - 1)
                progress.setLabelText(f"Loading {source_name}: {idx}/{len(candidates)}\n{path.name}")
                QApplication.processEvents()
            try:
                if typ == "REVIEW":
                    records.extend(review_rows_to_viewer_records(path))
                elif typ == "PSC":
                    _, _, recs = parse_psc_file_detail(path)
                    records.extend(recs)
                else:
                    records.extend(parse_file(path, typ))
            except Exception as e:
                records.append(LogRecord(None, typ, path.name, 0, "ERROR", "Viewer", f"Parse failed: {e}", ""))
        if progress:
            progress.setValue(len(candidates))
            progress.setLabelText(f"Filtering and sorting {len(records)} rows ...")
            QApplication.processEvents()

        # Apply the same date filter as merge, based on each row timestamp.
        # Important: this never uses file creation/modified time.
        start_date = parse_user_date(opt.start_date) if opt.use_start and opt.start_date else None
        end_date = parse_user_date(opt.end_date) if opt.use_end and opt.end_date else None
        if start_date or end_date:
            before = len(records)
            records = [r for r in records if in_date_range(r.timestamp, None, start_date, end_date)]
            self.log(f"Date filter kept {len(records)}/{before} rows for {source_name}.")
        if self.apply_noise_chk.isChecked():
            engine = NoiseEngine()
            before = len(records)
            filtered_records: list[LogRecord] = []
            matched_rule_ids: list[int] = []
            for r in records:
                matched_rule = engine.match_rule(r)
                if matched_rule is not None:
                    rid = matched_rule.get("id")
                    if rid is not None:
                        matched_rule_ids.append(int(rid))
                    continue
                filtered_records.append(r)
            records = filtered_records
            increment_noise_hits(matched_rule_ids)
            hidden = before - len(records)
            if hidden:
                self.log(f"Noise Filter hidden rows: {hidden}")
            else:
                self.log("Noise Filter: no rows matched approved enabled rules in this pane.")
        records.sort(key=lambda r: (r.timestamp is None, r.timestamp or datetime.max, r.source_type, r.filename, r.line_no))
        self.log(f"Loaded {len(records)} rows from {len(candidates)} files for {source_name}.")
        return records

    def load_side(self, side: str):
        source = self.left_source.currentText() if side == "left" else self.right_source.currentText()
        self.log(f"Loading {side}: {source} ...")
        progress = self.make_progress("Dual Log Viewer", f"Preparing {side}: {source} ...")
        progress.show()
        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            records = self.source_to_records(source, progress)
            if progress.wasCanceled():
                return
            progress.setLabelText(f"Building table model for {side}: {len(records)} rows ...")
            QApplication.processEvents()
            rows = [record_to_viewer_row(r) for r in records]
            ts_index = [(row["_ts"], i) for i, row in enumerate(rows) if isinstance(row.get("_ts"), datetime)]
            ts_index.sort(key=lambda x: x[0])
            file_hint = source
            if rows:
                files = sorted({str(r.get("File", "")) for r in rows if r.get("File", "")})
                if len(files) == 1:
                    file_hint = files[0]
                elif len(files) > 1:
                    file_hint = f"{len(files)} files loaded"
            if side == "left":
                self.left_all_rows = rows
                self.left_file_label.setText(f"File: {file_hint}")
            else:
                self.right_all_rows = rows
                self.right_file_label.setText(f"File: {file_hint}")
            self.apply_view_filters(side)
            self.log(f"{side.capitalize()} loaded: {len(rows)} rows")
        except Exception:
            write_startup_log("Dual Log Viewer load_side failed.\n\n" + traceback.format_exc())
            QMessageBox.critical(self, "Dual Log Viewer", "Viewer loading failed.\n\n" + traceback.format_exc() + f"\n\nLog:\n{startup_log_path()}")
        finally:
            QApplication.restoreOverrideCursor()
            progress.close()

    def apply_noise_rules_now(self):
        """Apply approved noise rules at the operator-selected timing."""
        if self.apply_noise_chk.isChecked():
            self.log("Applying approved Noise rules to both panes...")
        else:
            self.log("Noise filter disabled. Reloading both panes...")
        self.load_both()

    def find_next(self, side: str):
        table = self.left_table if side == "left" else self.right_table
        model = self.left_model if side == "left" else self.right_model
        edit = self.left_search if side == "left" else self.right_search
        term = edit.text().strip().lower()
        if not term:
            return
        current = -1
        indexes = table.selectionModel().selectedRows() if table.selectionModel() else []
        if indexes:
            current = indexes[0].row()
        total = model.rowCount()
        if total <= 0:
            return
        for step in range(1, total + 1):
            i = (current + step) % total
            row = model.row_at(i) or {}
            hay = " ".join(str(row.get(k, "")) for k in ("Timestamp", "SourceType", "File", "Level", "Category", "Message", "Raw")).lower()
            if term in hay:
                table.selectRow(i)
                table.scrollTo(model.index(i, 0), QTableView.PositionAtCenter)
                self.show_detail(side, row)
                self.log(f"Found in {side}: row {i + 1}/{total}")
                return
        self.log(f"No match in {side}: {term}")

    def row_selected(self, side: str):
        if self._syncing:
            return
        table = self.left_table if side == "left" else self.right_table
        model = self.left_model if side == "left" else self.right_model
        indexes = table.selectionModel().selectedRows()
        if not indexes:
            return
        row_idx = indexes[0].row()
        row = model.row_at(row_idx)
        if not row:
            return
        self.show_detail(side, row)
        ts = row.get("_ts")
        if not isinstance(ts, datetime):
            return
        self.jump_other_side(side, ts)

    def show_detail(self, side: str, row: dict[str, Any]):
        raw = row.get("Raw", "")
        msg = row.get("Message", "")
        self.detail.setPlainText(
            f"Selected {side}\n"
            f"Timestamp: {row.get('Timestamp','')}\n"
            f"SourceType: {row.get('SourceType','')}\n"
            f"File: {row.get('File','')}  Line: {row.get('Line','')}\n"
            f"Message: {msg}\n\nRaw:\n{raw}"
        )

    def jump_other_side(self, side: str, ts: datetime):
        other_ts = self.right_ts if side == "left" else self.left_ts
        other_table = self.right_table if side == "left" else self.left_table
        other_model = self.right_model if side == "left" else self.left_model
        if not other_ts:
            return
        # Binary search without importing bisect key dependency for older environments.
        lo, hi = 0, len(other_ts)
        while lo < hi:
            mid = (lo + hi) // 2
            if other_ts[mid][0] < ts:
                lo = mid + 1
            else:
                hi = mid
        candidates = []
        for pos in (lo - 1, lo, lo + 1):
            if 0 <= pos < len(other_ts):
                candidates.append(other_ts[pos])
        if not candidates:
            return
        nearest_ts, nearest_row = min(candidates, key=lambda x: abs((x[0] - ts).total_seconds()))
        delta = abs((nearest_ts - ts).total_seconds())
        tol = self.tolerance_spin.value()
        self._syncing = True
        try:
            other_table.selectRow(nearest_row)
            other_table.scrollTo(other_model.index(nearest_row, 0), QTableView.PositionAtCenter)
        finally:
            self._syncing = False
        msg = f"Cross reference: nearest row delta {delta:.3f} sec"
        if tol and delta > tol:
            msg += f" (outside ±{tol} sec)"
        self.log(msg)

    def selected_record_from_side(self, side: str) -> Optional[LogRecord]:
        table = self.left_table if side == "left" else self.right_table
        model = self.left_model if side == "left" else self.right_model
        indexes = table.selectionModel().selectedRows() if table.selectionModel() else []
        if not indexes:
            return None
        row = model.row_at(indexes[0].row())
        if not row:
            return None
        return LogRecord(row.get("_ts"), str(row.get("SourceType", "")), str(row.get("File", "")), int(row.get("Line", 0) or 0), str(row.get("Level", "")), str(row.get("Category", "")), str(row.get("Message", "")), str(row.get("Raw", "")))

    def copy_rule_text(self, side: str):
        rec = self.selected_record_from_side(side)
        if rec is None:
            QMessageBox.information(self, "Noise Learning", "Select a row first.")
            return
        text = _compact_pattern(rec.message or rec.raw)
        QApplication.clipboard().setText(text)
        self.log(f"Copied {side} rule text to clipboard: {text}")

    def approve_selected_noise(self, side: str):
        rec = self.selected_record_from_side(side)
        if rec is None:
            QMessageBox.information(self, "Noise Learning", "Select a row first.")
            return
        engine = NoiseEngine()
        suggested = _compact_pattern(rec.message or rec.raw)
        pattern, ok = QInputDialog.getText(self, "Approve Noise Rule", "Rule text to hide (contains):", text=suggested)
        if not ok or not pattern.strip():
            return
        # Store exactly what the operator approved. Matching is normalized,
        # so small whitespace differences will not prevent filtering.
        rid = engine.add_rule(pattern.strip(), (rec.source_type or "ANY").upper(), "contains", "hide", "both", "Approved by operator from Explorer selected row")
        engine.add_sample(rec, "noise-approved", 100)
        if self.noise_apply_mode.currentIndex() == 0:
            QMessageBox.information(self, "Noise Learning", f"Approved noise rule ID {rid}.\nApplying filter now.")
            self.apply_noise_rules_now()
        else:
            QMessageBox.information(self, "Noise Learning", f"Approved noise rule ID {rid}.\nThe rule is saved. Press Apply Rules Now when you want to apply it to the current Explorer view.")

    def mark_selected_not_noise(self):
        rec = self.selected_record_from_side("left") or self.selected_record_from_side("right")
        if rec is None:
            QMessageBox.information(self, "Noise Learning", "Select a row first.")
            return
        NoiseEngine().mark_not_noise(rec)
        QMessageBox.information(self, "Noise Learning", "Saved as Not Noise training sample. Approved rules were not changed.")

    def manage_noise_rules(self):
        dlg = NoiseRuleManagerDialog(self)
        dlg.exec()
        if self.noise_apply_mode.currentIndex() == 0:
            self.apply_noise_rules_now()
        else:
            self.log("Noise rules changed. Press Apply Rules Now to apply them to the current Explorer view.")

    def export_selected_pair(self):
        left_rows = self.left_table.selectionModel().selectedRows()
        right_rows = self.right_table.selectionModel().selectedRows()
        if not left_rows and not right_rows:
            QMessageBox.information(self, "Dual Log Viewer", "Select at least one row.")
            return
        out_dir = Path(self.parent_window.output_edit.text().strip() or self.parent_window.source_edit.text().strip() or str(Path.home()))
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"DualLogViewer_SelectedPair_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        with out_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Side"] + VIEWER_COLUMNS + ["Raw"])
            if left_rows:
                row = self.left_model.row_at(left_rows[0].row())
                if row:
                    writer.writerow(["Left"] + [row.get(c, "") for c in VIEWER_COLUMNS] + [row.get("Raw", "")])
            if right_rows:
                row = self.right_model.row_at(right_rows[0].row())
                if row:
                    writer.writerow(["Right"] + [row.get(c, "") for c in VIEWER_COLUMNS] + [row.get("Raw", "")])
        QMessageBox.information(self, "Dual Log Viewer", f"Exported.\n\n{out_path}")



# v34 Multi Log Viewer override ------------------------------------------------
# Keep the legacy DualLogViewer above as a fallback implementation, then replace
# it with a multi-pane Explorer. MainWindow still calls DualLogViewer(self), so
# this alias preserves compatibility while adding 1/2/3/4 pane modes.
LegacyDualLogViewer = DualLogViewer

class MultiPaneLogViewer(LegacyDualLogViewer):
    SOURCES = ["Merged", "WS", "CGA", "CSA", "MRSERVER", "GESYS", "LAIS", "PSC", "Review", "Custom File"]
    MAX_PANES = 4

    def __init__(self, parent_window: 'MainWindow'):
        QWidget.__init__(self)
        self.parent_window = parent_window
        self.setWindowTitle("Log Viewer / Log Explorer - 1 to 4 Logs")
        self.resize(1500, 820)
        self._syncing = False
        self.models: list[LogTableModel] = [LogTableModel([]) for _ in range(self.MAX_PANES)]
        self.tables: list[QTableView] = []
        self.sources: list[QComboBox] = []
        self.search_edits: list[QLineEdit] = []
        self.file_labels: list[QLabel] = []
        self.panes: list[QWidget] = []
        self.all_rows: list[list[dict[str, Any]]] = [[] for _ in range(self.MAX_PANES)]
        self.ts_indexes: list[list[tuple[datetime, int]]] = [[] for _ in range(self.MAX_PANES)]
        self.build_ui()
        self.update_view_mode()

    def side_index(self, side: Any) -> int:
        if isinstance(side, int):
            return max(0, min(self.MAX_PANES - 1, side))
        mapping = {"left": 0, "right": 1, "pane1": 0, "pane2": 1, "pane3": 2, "pane4": 3}
        return mapping.get(str(side).lower(), 0)

    def pane_name(self, idx: int) -> str:
        names = ["Left", "Right", "Log 3", "Log 4"]
        return names[idx] if 0 <= idx < len(names) else f"Log {idx + 1}"

    def visible_pane_count(self) -> int:
        text = self.mode_combo.currentText() if hasattr(self, "mode_combo") else "2 logs (Dual)"
        m = re.search(r"(\d+)", text)
        return max(1, min(self.MAX_PANES, int(m.group(1)) if m else 2))

    def build_ui(self):
        root = QVBoxLayout(self)

        top = QHBoxLayout()
        self.mode_combo = QComboBox()
        self.mode_combo.addItems(["1 log", "2 logs (Dual)", "3 logs", "4 logs"])
        self.mode_combo.setCurrentText("2 logs (Dual)")
        self.mode_combo.setToolTip("Switch Log Viewer layout between one, two, three, or four panes. Default is Dual mode.")
        self.tolerance_spin = QSpinBox(); self.tolerance_spin.setRange(0, 3600); self.tolerance_spin.setValue(10); self.tolerance_spin.setSuffix(" sec")
        self.load_visible_btn = QPushButton("Load Visible Logs")
        self.export_pair_btn = QPushButton("Export Selected Rows")
        self.apply_noise_chk = QCheckBox("Apply Noise Filter")
        self.apply_noise_chk.setChecked(True)
        self.noise_apply_mode = QComboBox(); self.noise_apply_mode.addItems(["Apply immediately after adding rule", "Apply only when Apply button is pressed"])
        self.apply_rules_btn = QPushButton("Apply Rules Now")
        self.not_noise_btn = QPushButton("Mark Selected Not Noise")
        self.manage_noise_btn = QPushButton("Manage Rules")
        self.status = QLabel("Ready")
        top.addWidget(QLabel("View mode:")); top.addWidget(self.mode_combo)
        top.addSpacing(12)
        top.addWidget(QLabel("Jump tolerance:")); top.addWidget(self.tolerance_spin)
        top.addWidget(self.load_visible_btn)
        top.addWidget(self.export_pair_btn)
        top.addWidget(self.apply_noise_chk)
        top.addWidget(self.noise_apply_mode)
        top.addWidget(self.apply_rules_btn)
        top.addWidget(self.not_noise_btn)
        top.addWidget(self.manage_noise_btn)
        top.addStretch(1)
        root.addLayout(top)

        time_bar = QHBoxLayout()
        self.viewer_time_enable = QCheckBox("Enable Viewer Time Range")
        self.viewer_start_edit = QLineEdit(); self.viewer_start_edit.setPlaceholderText("Start: 2026/07/04 13:07:58.6")
        self.viewer_end_edit = QLineEdit(); self.viewer_end_edit.setPlaceholderText("End: 2026/07/04 13:30:00.0")
        self.apply_time_current_btn = QPushButton("Apply to Focused")
        self.apply_time_all_btn = QPushButton("Apply to Visible")
        self.clear_time_btn = QPushButton("Clear Time Range")
        self.set_start_btn = QPushButton("Set Start from Selected")
        self.set_end_btn = QPushButton("Set End from Selected")
        time_bar.addWidget(self.viewer_time_enable)
        time_bar.addWidget(self.viewer_start_edit, 1)
        time_bar.addWidget(self.viewer_end_edit, 1)
        time_bar.addWidget(self.apply_time_current_btn)
        time_bar.addWidget(self.apply_time_all_btn)
        time_bar.addWidget(self.clear_time_btn)
        time_bar.addWidget(self.set_start_btn)
        time_bar.addWidget(self.set_end_btn)
        root.addLayout(time_bar)

        self.main_splitter = QSplitter(Qt.Horizontal)
        self.main_splitter.setHandleWidth(10)
        self.main_splitter.setOpaqueResize(True)
        self.main_splitter.setChildrenCollapsible(False)
        self.main_splitter.setToolTip("Drag splitters to resize each log pane. Timestamp stays fixed and Message expands/shrinks.")

        pane_colors = ["#EAF3FF", "#EAF8EA", "#FFF4DF", "#F2EAFF"]
        default_sources = ["Merged", "PSC", "Review", "MRSERVER"]
        for i in range(self.MAX_PANES):
            pane = QWidget()
            layout = QVBoxLayout(pane); layout.setContentsMargins(4, 4, 4, 4)
            title = QLabel(self.pane_name(i).upper())
            title.setStyleSheet(f"font-weight:bold; background:{pane_colors[i]}; padding:4px;")
            layout.addWidget(title)

            source = QComboBox(); source.addItems(self.SOURCES); source.setCurrentText(default_sources[i])
            source.setToolTip(f"Select the log type or Custom File shown in {self.pane_name(i)} pane.")
            load_btn = QPushButton("Load This")
            load_btn.setToolTip(f"Load only {self.pane_name(i)} pane. Use Load Visible Logs to load all visible panes at once.")
            ctrl = QHBoxLayout()
            ctrl.addWidget(QLabel(f"{self.pane_name(i)} Log:")); ctrl.addWidget(source, 1); ctrl.addWidget(load_btn)
            layout.addLayout(ctrl)

            search = QLineEdit(); search.setPlaceholderText(f"Search {self.pane_name(i)} message/raw...")
            search.setToolTip(f"Search only {self.pane_name(i)} pane. It searches Timestamp, Message, Raw, File, Level and Category.")
            find_btn = QPushButton("Find")
            search_layout = QHBoxLayout()
            search_layout.addWidget(QLabel("Search:")); search_layout.addWidget(search, 1); search_layout.addWidget(find_btn)
            layout.addLayout(search_layout)

            file_label = QLabel("File: not loaded")
            file_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
            layout.addWidget(file_label)

            noise_bar = QHBoxLayout()
            add_noise = QPushButton("Add to Noise Rule")
            copy_rule = QPushButton("Copy Rule Text")
            add_noise.setToolTip(f"Create a Noise rule from the selected {self.pane_name(i)} row.")
            copy_rule.setToolTip(f"Copy a compact rule candidate from the selected {self.pane_name(i)} row.")
            noise_bar.addWidget(add_noise); noise_bar.addWidget(copy_rule); noise_bar.addStretch(1)
            layout.addLayout(noise_bar)

            table = QTableView(); table.setModel(self.models[i])
            table.setSelectionBehavior(QTableView.SelectRows)
            table.setSelectionMode(QTableView.SingleSelection)
            table.setAlternatingRowColors(True)
            table.setSortingEnabled(False)
            table.verticalHeader().setDefaultSectionSize(20)
            table.verticalHeader().setVisible(False)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            table.setColumnWidth(0, 170)
            table.setColumnWidth(1, 620)
            table.setColumnWidth(2, 90)
            table.setColumnWidth(3, 200)
            table.setColumnWidth(4, 60)
            table.setColumnWidth(5, 70)
            table.setColumnWidth(6, 100)
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            layout.addWidget(table, 1)

            self.main_splitter.addWidget(pane)
            self.panes.append(pane)
            self.sources.append(source)
            self.search_edits.append(search)
            self.file_labels.append(file_label)
            self.tables.append(table)

            load_btn.clicked.connect(lambda _=False, idx=i: self.load_pane(idx))
            find_btn.clicked.connect(lambda _=False, idx=i: self.find_next(idx))
            search.returnPressed.connect(lambda idx=i: self.find_next(idx))
            add_noise.clicked.connect(lambda _=False, idx=i: self.approve_selected_noise(idx))
            copy_rule.clicked.connect(lambda _=False, idx=i: self.copy_rule_text(idx))
            table.selectionModel().selectionChanged.connect(lambda *_args, idx=i: self.row_selected(idx))

        root.addWidget(self.main_splitter, 1)
        self.detail = QTextEdit(); self.detail.setReadOnly(True); self.detail.setMaximumHeight(150)
        root.addWidget(self.detail)
        root.addWidget(self.status)

        self.mode_combo.currentIndexChanged.connect(self.update_view_mode)
        self.load_visible_btn.clicked.connect(self.load_visible)
        self.export_pair_btn.clicked.connect(self.export_selected_pair)
        self.apply_noise_chk.toggled.connect(lambda _=False: self.apply_noise_rules_now())
        self.apply_rules_btn.clicked.connect(self.apply_noise_rules_now)
        self.not_noise_btn.clicked.connect(self.mark_selected_not_noise)
        self.manage_noise_btn.clicked.connect(self.manage_noise_rules)
        self.apply_time_current_btn.clicked.connect(self.apply_time_to_focused)
        self.apply_time_all_btn.clicked.connect(self.apply_time_to_visible)
        self.clear_time_btn.clicked.connect(self.clear_viewer_time_range)
        self.set_start_btn.clicked.connect(lambda: self.set_time_from_selected("start"))
        self.set_end_btn.clicked.connect(lambda: self.set_time_from_selected("end"))

        # Backward-compatible aliases used by inherited helper code and older tests.
        self.left_model = self.models[0]; self.right_model = self.models[1]
        self.left_table = self.tables[0]; self.right_table = self.tables[1]
        self.left_source = self.sources[0]; self.right_source = self.sources[1]
        self.left_search = self.search_edits[0]; self.right_search = self.search_edits[1]
        self.left_file_label = self.file_labels[0]; self.right_file_label = self.file_labels[1]
        self.left_all_rows = self.all_rows[0]; self.right_all_rows = self.all_rows[1]
        self.left_ts = self.ts_indexes[0]; self.right_ts = self.ts_indexes[1]

    def log(self, text: str):
        self.status.setText(text)
        try:
            self.parent_window.log("[Viewer] " + text)
        except Exception:
            pass

    def visible_indices(self) -> list[int]:
        return list(range(self.visible_pane_count()))

    def update_view_mode(self):
        count = self.visible_pane_count()
        for i, pane in enumerate(self.panes):
            pane.setVisible(i < count)
        self.main_splitter.setSizes([max(1, 1000 // count) if i < count else 0 for i in range(self.MAX_PANES)])
        self.log(f"Viewer mode changed: {count} log pane(s). Default is Dual mode.")

    def load_visible(self):
        for idx in self.visible_indices():
            self.load_pane(idx)

    def load_both(self):
        self.load_visible()

    def current_viewer_time_range(self) -> tuple[Optional[datetime], Optional[datetime]]:
        if not self.viewer_time_enable.isChecked():
            return None, None
        start = parse_viewer_datetime_text(self.viewer_start_edit.text(), is_end=False) if self.viewer_start_edit.text().strip() else None
        end = parse_viewer_datetime_text(self.viewer_end_edit.text(), is_end=True) if self.viewer_end_edit.text().strip() else None
        if start and end and start > end:
            raise ValueError("Viewer Start time is later than End time.")
        return start, end

    def apply_view_filters(self, side: Any):
        idx = self.side_index(side)
        try:
            start, end = self.current_viewer_time_range()
        except Exception as e:
            QMessageBox.warning(self, "Viewer Time Range", str(e))
            return
        base_rows = self.all_rows[idx]
        if start or end:
            rows = []
            for row in base_rows:
                ts = row.get("_ts")
                if not isinstance(ts, datetime):
                    continue
                if start and ts < start:
                    continue
                if end and ts > end:
                    continue
                rows.append(row)
        else:
            rows = list(base_rows)
        ts_index = [(row["_ts"], i) for i, row in enumerate(rows) if isinstance(row.get("_ts"), datetime)]
        ts_index.sort(key=lambda x: x[0])
        self.models[idx].set_rows(rows)
        self.ts_indexes[idx] = ts_index
        self._sync_aliases()
        label = self.pane_name(idx)
        if start or end:
            self.log(f"Viewer time filter {label}: {len(rows)}/{len(base_rows)} rows")
        else:
            self.log(f"Viewer time filter cleared {label}: {len(rows)} rows")

    def _sync_aliases(self):
        self.left_all_rows = self.all_rows[0]; self.right_all_rows = self.all_rows[1]
        self.left_ts = self.ts_indexes[0]; self.right_ts = self.ts_indexes[1]

    def apply_time_to_visible(self):
        for idx in self.visible_indices():
            self.apply_view_filters(idx)

    def apply_time_to_focused(self):
        for idx in self.visible_indices():
            if self.tables[idx].hasFocus() or self.search_edits[idx].hasFocus():
                self.apply_view_filters(idx)
                return
        self.apply_time_to_visible()

    def clear_viewer_time_range(self):
        self.viewer_time_enable.setChecked(False)
        self.viewer_start_edit.clear()
        self.viewer_end_edit.clear()
        self.apply_time_to_visible()

    def set_time_from_selected(self, target: str):
        for idx in self.visible_indices():
            indexes = self.tables[idx].selectionModel().selectedRows()
            if indexes:
                row = self.models[idx].row_at(indexes[0].row()) or {}
                ts_text = row.get("Timestamp", "")
                if ts_text:
                    if target == "start":
                        self.viewer_start_edit.setText(str(ts_text))
                    else:
                        self.viewer_end_edit.setText(str(ts_text))
                    self.viewer_time_enable.setChecked(True)
                    self.log(f"Set Viewer {target} from selected {self.pane_name(idx)}: {ts_text}")
                    return
        QMessageBox.information(self, "Viewer Time Range", "Select a row first.")

    def load_pane(self, side: Any):
        idx = self.side_index(side)
        source = self.sources[idx].currentText()
        label = self.pane_name(idx)
        self.log(f"Loading {label}: {source} ...")
        progress = self.make_progress("Log Viewer", f"Preparing {label}: {source} ...")
        progress.show()
        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            records = self.source_to_records(source, progress)
            if progress.wasCanceled():
                return
            progress.setLabelText(f"Building table model for {label}: {len(records)} rows ...")
            QApplication.processEvents()
            rows = [record_to_viewer_row(r) for r in records]
            file_hint = source
            if rows:
                files = sorted({str(r.get("File", "")) for r in rows if r.get("File", "")})
                if len(files) == 1:
                    file_hint = files[0]
                elif len(files) > 1:
                    file_hint = f"{len(files)} files loaded"
            self.all_rows[idx] = rows
            self.file_labels[idx].setText(f"File: {file_hint}")
            self.apply_view_filters(idx)
            self.log(f"{label} loaded: {len(rows)} rows")
        except Exception:
            write_startup_log("Multi Log Viewer load_pane failed.\n\n" + traceback.format_exc())
            QMessageBox.critical(self, "Log Viewer", "Viewer loading failed.\n\n" + traceback.format_exc() + f"\n\nLog:\n{startup_log_path()}")
        finally:
            QApplication.restoreOverrideCursor()
            progress.close()

    def load_side(self, side: str):
        self.load_pane(side)

    def apply_noise_rules_now(self):
        if self.apply_noise_chk.isChecked():
            self.log("Applying approved Noise rules to visible panes...")
        else:
            self.log("Noise filter disabled. Reloading visible panes...")
        self.load_visible()

    def find_next(self, side: Any):
        idx = self.side_index(side)
        table = self.tables[idx]
        model = self.models[idx]
        edit = self.search_edits[idx]
        term = edit.text().strip().lower()
        if not term:
            return
        current = -1
        indexes = table.selectionModel().selectedRows() if table.selectionModel() else []
        if indexes:
            current = indexes[0].row()
        total = model.rowCount()
        if total <= 0:
            return
        for step in range(1, total + 1):
            i = (current + step) % total
            row = model.row_at(i) or {}
            hay = " ".join(str(row.get(k, "")) for k in ("Timestamp", "SourceType", "File", "Level", "Category", "Message", "Raw")).lower()
            if term in hay:
                table.selectRow(i)
                table.scrollTo(model.index(i, 0), QTableView.PositionAtCenter)
                self.show_detail(idx, row)
                self.log(f"Found in {self.pane_name(idx)}: row {i + 1}/{total}")
                return
        self.log(f"No match in {self.pane_name(idx)}: {term}")

    def row_selected(self, side: Any):
        if self._syncing:
            return
        idx = self.side_index(side)
        table = self.tables[idx]
        model = self.models[idx]
        indexes = table.selectionModel().selectedRows()
        if not indexes:
            return
        row = model.row_at(indexes[0].row())
        if not row:
            return
        self.show_detail(idx, row)
        ts = row.get("_ts")
        if isinstance(ts, datetime):
            self.jump_other_panes(idx, ts)

    def show_detail(self, side: Any, row: dict[str, Any]):
        idx = self.side_index(side)
        raw = row.get("Raw", "")
        msg = row.get("Message", "")
        self.detail.setPlainText(
            f"Selected {self.pane_name(idx)}\n"
            f"Timestamp: {row.get('Timestamp','')}\n"
            f"SourceType: {row.get('SourceType','')}\n"
            f"File: {row.get('File','')}  Line: {row.get('Line','')}\n"
            f"Message: {msg}\n\nRaw:\n{raw}"
        )

    def jump_other_panes(self, source_idx: int, ts: datetime):
        tol = self.tolerance_spin.value()
        messages = []
        self._syncing = True
        try:
            for idx in self.visible_indices():
                if idx == source_idx:
                    continue
                ts_index = self.ts_indexes[idx]
                if not ts_index:
                    continue
                lo, hi = 0, len(ts_index)
                while lo < hi:
                    mid = (lo + hi) // 2
                    if ts_index[mid][0] < ts:
                        lo = mid + 1
                    else:
                        hi = mid
                candidates = []
                for pos in (lo - 1, lo, lo + 1):
                    if 0 <= pos < len(ts_index):
                        candidates.append(ts_index[pos])
                if not candidates:
                    continue
                nearest_ts, nearest_row = min(candidates, key=lambda x: abs((x[0] - ts).total_seconds()))
                delta = abs((nearest_ts - ts).total_seconds())
                self.tables[idx].selectRow(nearest_row)
                self.tables[idx].scrollTo(self.models[idx].index(nearest_row, 0), QTableView.PositionAtCenter)
                status = f"{self.pane_name(idx)} Δ{delta:.3f}s"
                if tol and delta > tol:
                    status += f" outside ±{tol}s"
                messages.append(status)
        finally:
            self._syncing = False
        if messages:
            self.log("Cross reference: " + "; ".join(messages))

    def jump_other_side(self, side: str, ts: datetime):
        self.jump_other_panes(self.side_index(side), ts)

    def selected_record_from_pane(self, idx: int) -> Optional[LogRecord]:
        table = self.tables[idx]
        model = self.models[idx]
        indexes = table.selectionModel().selectedRows() if table.selectionModel() else []
        if not indexes:
            return None
        row = model.row_at(indexes[0].row())
        if not row:
            return None
        return LogRecord(row.get("_ts"), str(row.get("SourceType", "")), str(row.get("File", "")), int(row.get("Line", 0) or 0), str(row.get("Level", "")), str(row.get("Category", "")), str(row.get("Message", "")), str(row.get("Raw", "")))

    def selected_record_from_side(self, side: str) -> Optional[LogRecord]:
        return self.selected_record_from_pane(self.side_index(side))

    def selected_any_record(self) -> Optional[LogRecord]:
        for idx in self.visible_indices():
            rec = self.selected_record_from_pane(idx)
            if rec is not None:
                return rec
        return None

    def copy_rule_text(self, side: Any):
        idx = self.side_index(side)
        rec = self.selected_record_from_pane(idx)
        if rec is None:
            QMessageBox.information(self, "Noise Learning", "Select a row first.")
            return
        text = _compact_pattern(rec.message or rec.raw)
        QApplication.clipboard().setText(text)
        self.log(f"Copied {self.pane_name(idx)} rule text to clipboard: {text}")

    def approve_selected_noise(self, side: Any):
        idx = self.side_index(side)
        rec = self.selected_record_from_pane(idx)
        if rec is None:
            QMessageBox.information(self, "Noise Learning", "Select a row first.")
            return
        engine = NoiseEngine()
        suggested = _compact_pattern(rec.message or rec.raw)
        pattern, ok = QInputDialog.getText(self, "Approve Noise Rule", "Rule text to hide (contains):", text=suggested)
        if not ok or not pattern.strip():
            return
        rid = engine.add_rule(pattern.strip(), (rec.source_type or "ANY").upper(), "contains", "hide", "both", "Approved by operator from Explorer selected row")
        engine.add_sample(rec, "noise-approved", 100)
        if self.noise_apply_mode.currentIndex() == 0:
            QMessageBox.information(self, "Noise Learning", f"Approved noise rule ID {rid}.\nApplying filter now.")
            self.apply_noise_rules_now()
        else:
            QMessageBox.information(self, "Noise Learning", f"Approved noise rule ID {rid}.\nThe rule is saved. Press Apply Rules Now when you want to apply it to the current Explorer view.")

    def mark_selected_not_noise(self):
        rec = self.selected_any_record()
        if rec is None:
            QMessageBox.information(self, "Noise Learning", "Select a row first.")
            return
        NoiseEngine().mark_not_noise(rec)
        QMessageBox.information(self, "Noise Learning", "Saved as Not Noise training sample. Approved rules were not changed.")

    def manage_noise_rules(self):
        dlg = NoiseRuleManagerDialog(self)
        dlg.exec()
        if self.noise_apply_mode.currentIndex() == 0:
            self.apply_noise_rules_now()
        else:
            self.log("Noise rules changed. Press Apply Rules Now to apply them to the current Explorer view.")

    def export_selected_pair(self):
        out_dir = Path(self.parent_window.output_edit.text().strip() or self.parent_window.source_edit.text().strip() or str(Path.home()))
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"LogViewer_SelectedRows_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        selected = []
        for idx in self.visible_indices():
            rows = self.tables[idx].selectionModel().selectedRows()
            if rows:
                row = self.models[idx].row_at(rows[0].row())
                if row:
                    selected.append((self.pane_name(idx), row, self.models[idx].columns))
        if not selected:
            QMessageBox.information(self, "Log Viewer", "Select at least one row.")
            return
        with out_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Pane"] + VIEWER_COLUMNS + ["Raw"])
            for pane, row, _cols in selected:
                writer.writerow([pane] + [row.get(c, "") for c in VIEWER_COLUMNS] + [row.get("Raw", "")])
        QMessageBox.information(self, "Log Viewer", f"Exported.\n\n{out_path}")

# Replace the legacy class name used by MainWindow.
DualLogViewer = MultiPaneLogViewer



@dataclass
class DiscoveredFile:
    path: str
    log_type: str
    row_count: int
    start_ts: Optional[datetime]
    end_ts: Optional[datetime]
    size_bytes: int
    error: str = ""


def discover_one_file(path: Path, include_unknown: bool) -> DiscoveredFile:
    """Parse one candidate file for Smart File Discovery.

    This intentionally uses the same parser/timestamp logic as Merge so the
    preflight list and the final merge cannot disagree about date filtering.
    File created/modified timestamps are never used.
    """
    typ = "REVIEW" if is_review_file(path.name) else classify_file(path, include_unknown)
    if typ is None:
        return DiscoveredFile(str(path), "", 0, None, None, path.stat().st_size if path.exists() else 0)
    try:
        if typ == "REVIEW":
            rows = parse_review_file(path)
            timestamps = []
            for r in rows:
                text_dt = " ".join(str(r.get(k, "")) for k in ("date", "time") if r.get(k, ""))
                ts = parse_datetime_from_text(text_dt) if text_dt.strip() else None
                if ts:
                    timestamps.append(ts)
            return DiscoveredFile(str(path), typ, len(rows), min(timestamps) if timestamps else None, max(timestamps) if timestamps else None, path.stat().st_size)
        if typ == "PSC":
            psc_log, psc_params, parsed = parse_psc_file_detail(path)
            timestamps = [r.timestamp for r in parsed if r.timestamp]
            return DiscoveredFile(str(path), typ, len(parsed) + len(psc_log) + len(psc_params), min(timestamps) if timestamps else None, max(timestamps) if timestamps else None, path.stat().st_size)
        parsed = parse_file(path, typ)
        timestamps = [r.timestamp for r in parsed if r.timestamp]
        return DiscoveredFile(str(path), typ, len(parsed), min(timestamps) if timestamps else None, max(timestamps) if timestamps else None, path.stat().st_size)
    except Exception as exc:
        return DiscoveredFile(str(path), typ or "UNKNOWN", 0, None, None, path.stat().st_size if path.exists() else 0, str(exc))


class SmartFileDiscoveryDialog(QDialog):
    def __init__(self, parent, base_options: RunOptions):
        super().__init__(parent)
        self.setWindowTitle("Smart File Discovery")
        self.resize(1050, 720)
        self.base_options = base_options
        self.discovered: list[DiscoveredFile] = []
        self.filtered: list[DiscoveredFile] = []
        self.selected_files: list[str] = []
        self.selected_types: set[str] = set()
        self._build_ui()
        self._scan()

    def _build_ui(self):
        root = QVBoxLayout(self)
        info = QLabel("1) Detect log types in the selected folder. 2) Select log types and date range. 3) Confirm target files before starting.")
        info.setWordWrap(True)
        root.addWidget(info)

        type_box = QGroupBox("Detected Log Types")
        self.type_layout = QHBoxLayout(type_box)
        self.type_checks: dict[str, QCheckBox] = {}
        root.addWidget(type_box)

        range_box = QGroupBox("Date Range Filter (based on each file's parsed row timestamps, not file created/modified date)")
        rg = QHBoxLayout(range_box)
        self.enable_start = QCheckBox("Start")
        self.start_dt = QLineEdit()
        self.start_dt.setPlaceholderText("YYYY/MM/DD HH:MM:SS or YYYY-MM-DD")
        self.enable_end = QCheckBox("End")
        self.end_dt = QLineEdit()
        self.end_dt.setPlaceholderText("YYYY/MM/DD HH:MM:SS or YYYY-MM-DD")
        self.apply_btn = QPushButton("Apply Type/Date Filter")
        self.apply_btn.clicked.connect(self.apply_filter)
        for w in [self.enable_start, self.start_dt, self.enable_end, self.end_dt, self.apply_btn]:
            rg.addWidget(w)
        root.addWidget(range_box)

        self.summary_label = QLabel("Ready")
        root.addWidget(self.summary_label)
        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels(["Use", "Type", "Start", "End", "Rows", "Size MB", "File", "Error"])
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.Stretch)
        self.table.setSortingEnabled(True)
        root.addWidget(self.table, stretch=1)

        btns = QHBoxLayout()
        self.select_all_btn = QPushButton("Select All")
        self.select_all_btn.clicked.connect(lambda: self.set_all(True))
        self.clear_all_btn = QPushButton("Clear All")
        self.clear_all_btn.clicked.connect(lambda: self.set_all(False))
        self.start_btn = QPushButton("Start with Checked Files")
        self.start_btn.clicked.connect(self.accept_with_selection)
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.reject)
        for b in [self.select_all_btn, self.clear_all_btn, self.start_btn, self.cancel_btn]:
            btns.addWidget(b)
        root.addLayout(btns)

    def _scan(self):
        src = Path(self.base_options.source_folder)
        all_files = list(iter_files(src, self.base_options.recursive))
        progress = QProgressDialog("Scanning and parsing candidate files...", "Cancel", 0, max(1, len(all_files)), self)
        progress.setWindowTitle("Smart File Discovery")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(300)
        include_unknown = self.base_options.include_unknown
        discovered = []
        for i, path in enumerate(all_files, start=1):
            if progress.wasCanceled():
                break
            progress.setValue(i - 1)
            progress.setLabelText(f"Scanning {i}/{len(all_files)}\n{path.name}")
            QApplication.processEvents()
            meta = discover_one_file(path, include_unknown)
            if meta.log_type:
                discovered.append(meta)
        progress.setValue(len(all_files))
        self.discovered = discovered
        self.populate_type_checks()
        self.apply_filter()

    def populate_type_checks(self):
        while self.type_layout.count():
            item = self.type_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self.type_checks.clear()
        counts: dict[str, int] = {}
        for d in self.discovered:
            counts[d.log_type] = counts.get(d.log_type, 0) + 1
        preferred = ["WS", "CGA", "CSA", "MRSERVER", "GESYS", "LAIS", "PSC", "REVIEW", "UNKNOWN"]
        for typ in [t for t in preferred if t in counts] + sorted(t for t in counts if t not in preferred):
            chk = QCheckBox(f"{typ} ({counts[typ]})")
            chk.setChecked(False)
            chk.setToolTip(f"Include/exclude all detected {typ} files before final file confirmation.")
            chk.stateChanged.connect(self.apply_filter)
            self.type_checks[typ] = chk
            self.type_layout.addWidget(chk)
        self.type_layout.addStretch(1)

    def _range(self) -> tuple[Optional[datetime], Optional[datetime]]:
        start = end = None
        if self.enable_start.isChecked() and self.start_dt.text().strip():
            start = parse_viewer_datetime_text(self.start_dt.text().strip(), is_end=False)
        if self.enable_end.isChecked() and self.end_dt.text().strip():
            end = parse_viewer_datetime_text(self.end_dt.text().strip(), is_end=True)
        if start and end and start > end:
            raise ValueError("Start is later than End.")
        return start, end

    def _overlaps_range(self, d: DiscoveredFile, start: Optional[datetime], end: Optional[datetime]) -> bool:
        if not start and not end:
            return True
        # Files without parsed timestamps are kept for operator decision, because
        # they may still be important or require parser improvement.
        if not d.start_ts and not d.end_ts:
            return True
        s = d.start_ts or d.end_ts
        e = d.end_ts or d.start_ts
        if start and e and e < start:
            return False
        if end and s and s > end:
            return False
        return True

    def apply_filter(self):
        try:
            start, end = self._range()
        except Exception as exc:
            QMessageBox.warning(self, "Smart File Discovery", str(exc))
            return
        enabled = {typ for typ, chk in self.type_checks.items() if chk.isChecked()}
        self.selected_types = set(enabled)
        rows = [d for d in self.discovered if d.log_type in enabled and self._overlaps_range(d, start, end)]
        rows.sort(key=lambda d: (d.start_ts or datetime.max, d.log_type, Path(d.path).name))
        self.filtered = rows
        self.populate_file_table(rows)
        total_rows = sum(d.row_count for d in rows)
        total_mb = sum(d.size_bytes for d in rows) / (1024*1024)
        full_start = min((d.start_ts for d in self.discovered if d.start_ts), default=None)
        full_end = max((d.end_ts for d in self.discovered if d.end_ts), default=None)
        range_text = f"Detected range: {format_viewer_timestamp(full_start)} - {format_viewer_timestamp(full_end)}" if full_start or full_end else "Detected range: unknown"
        self.summary_label.setText(f"{range_text} | Filtered files: {len(rows)} | Estimated rows: {total_rows:,} | Size: {total_mb:.1f} MB")

    def populate_file_table(self, rows: list[DiscoveredFile]):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(rows))
        for r, d in enumerate(rows):
            chk = QTableWidgetItem("")
            chk.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            chk.setCheckState(Qt.Checked)
            chk.setData(Qt.UserRole, d.path)
            self.table.setItem(r, 0, chk)
            vals = [d.log_type, format_viewer_timestamp(d.start_ts), format_viewer_timestamp(d.end_ts), f"{d.row_count:,}", f"{d.size_bytes/(1024*1024):.1f}", str(Path(d.path).name), d.error]
            for c, v in enumerate(vals, start=1):
                item = QTableWidgetItem(str(v))
                item.setToolTip(d.path if c == 6 else str(v))
                self.table.setItem(r, c, item)
        self.table.setSortingEnabled(True)

    def set_all(self, checked: bool):
        for r in range(self.table.rowCount()):
            item = self.table.item(r, 0)
            if item:
                item.setCheckState(Qt.Checked if checked else Qt.Unchecked)

    def accept_with_selection(self):
        selected = []
        for r in range(self.table.rowCount()):
            item = self.table.item(r, 0)
            if item and item.checkState() == Qt.Checked:
                selected.append(str(item.data(Qt.UserRole)))
        if not selected:
            QMessageBox.warning(self, "Smart File Discovery", "No file is checked.")
            return
        self.selected_files = selected
        self.accept()

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.settings = QSettings("LogMerge", "NoExcel")
        self.worker: Optional[MergeWorker] = None
        self.progress_dialog = None
        self.import_worker = None
        self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")
        self.resize(860, 680)
        self.build_ui()
        self.load_settings()

    def build_ui(self):
        root = QVBoxLayout(self)

        title = QLabel(f"{APP_TITLE}")
        title.setStyleSheet("font-size: 20px; font-weight: bold;")
        root.addWidget(title)
        subtitle = QLabel("No Microsoft Excel / No VBA / No COM. v36: Plugin Manager + Update File Type ZIP.")
        root.addWidget(subtitle)

        folder_box = QGroupBox("Folders")
        folder_layout = QFormLayout(folder_box)
        self.source_edit = QLineEdit()
        self.source_btn = QPushButton("Browse...")
        self.source_btn.clicked.connect(self.pick_source)
        h1 = QHBoxLayout(); h1.addWidget(self.source_edit); h1.addWidget(self.source_btn)
        folder_layout.addRow("Source Log Folder", h1)
        self.output_edit = QLineEdit()
        self.output_btn = QPushButton("Browse...")
        self.output_btn.clicked.connect(self.pick_output)
        h2 = QHBoxLayout(); h2.addWidget(self.output_edit); h2.addWidget(self.output_btn)
        folder_layout.addRow("Output Folder", h2)
        self.recursive_chk = QCheckBox("Include all subfolders recursively")
        self.recursive_chk.setChecked(True)
        folder_layout.addRow("", self.recursive_chk)
        root.addWidget(folder_box)

        date_box = QGroupBox("Date Range")
        date_layout = QFormLayout(date_box)
        self.use_start = QCheckBox("Use Start Date")
        self.use_start.setChecked(False)
        self.start_date = QDateEdit(QDate.currentDate())
        self.start_date.setCalendarPopup(True)
        hs = QHBoxLayout(); hs.addWidget(self.use_start); hs.addWidget(self.start_date)
        date_layout.addRow("Start", hs)
        self.use_end = QCheckBox("Use End Date")
        self.use_end.setChecked(False)
        self.end_date = QDateEdit(QDate.currentDate())
        self.end_date.setCalendarPopup(True)
        he = QHBoxLayout(); he.addWidget(self.use_end); he.addWidget(self.end_date)
        date_layout.addRow("End", he)
        root.addWidget(date_box)

        meta_box = QGroupBox("Site Information")
        meta_layout = QFormLayout(meta_box)
        self.serial_combo = QComboBox(); self.serial_combo.setEditable(True)
        self.site_combo = QComboBox(); self.site_combo.setEditable(True)
        self.site_map = load_json_file("site_serial_map.json", DEFAULT_SITE_MAP)
        self.serial_combo.addItem(""); self.site_combo.addItem("")
        for item in self.site_map:
            self.serial_combo.addItem(str(item.get("serial", "")))
            self.site_combo.addItem(str(item.get("site", "")))
        self.serial_combo.currentTextChanged.connect(self.sync_site_from_serial)
        self.site_combo.currentTextChanged.connect(self.sync_serial_from_site)
        meta_layout.addRow("Serial Number", self.serial_combo)
        meta_layout.addRow("Site Name", self.site_combo)
        root.addWidget(meta_box)

        type_box = QGroupBox("Log Types")
        type_layout = QHBoxLayout(type_box)
        self.type_layout = type_layout
        self.plugin_checks: dict[str, QCheckBox] = {}
        self.chk_ws = QCheckBox("WS"); self.chk_ws.setChecked(False)
        self.chk_cga = QCheckBox("CGA"); self.chk_cga.setChecked(False)
        self.chk_csa = QCheckBox("CSA"); self.chk_csa.setChecked(False)
        self.chk_mrserver = QCheckBox("MRSERVER"); self.chk_mrserver.setChecked(False)
        self.chk_gesys = QCheckBox("GESYS"); self.chk_gesys.setChecked(False)
        self.chk_lais = QCheckBox("LAIS"); self.chk_lais.setChecked(False)
        self.chk_psc = QCheckBox("PSC"); self.chk_psc.setChecked(False)
        self.chk_review = QCheckBox("Review (Import Only)"); self.chk_review.setChecked(False)
        self.chk_review.setEnabled(False)
        self.chk_review.setToolTip("Review files are not merged. Use the Import Review Only button. Target: review.out, review.out.ar, review.out.*")
        self.chk_unknown = QCheckBox("UNKNOWN .log/.txt"); self.chk_unknown.setChecked(False)
        for w in [self.chk_ws, self.chk_cga, self.chk_csa, self.chk_mrserver, self.chk_gesys, self.chk_lais, self.chk_psc, self.chk_review, self.chk_unknown]:
            type_layout.addWidget(w)
        self.refresh_plugin_checkboxes()
        root.addWidget(type_box)

        perf_box = QGroupBox("Performance")
        perf_layout = QHBoxLayout(perf_box)
        self.chk_turbo_csv = QCheckBox("v8 Turbo CSV for large data")
        self.chk_turbo_csv.setChecked(True)
        self.chk_summary_xlsx = QCheckBox("CSV + Summary xlsx for large data")
        self.chk_summary_xlsx.setChecked(True)
        self.chk_fast_xlsx = QCheckBox("Fast xlsx mode")
        self.chk_fast_xlsx.setChecked(True)
        self.chk_skip_merged1 = QCheckBox("Skip Merged_1 duplicate")
        self.chk_skip_merged1.setChecked(True)
        self.chk_timestamp_text = QCheckBox("Timestamp as text for speed")
        self.chk_timestamp_text.setChecked(True)
        self.chk_disable_styles = QCheckBox("No styles for large xlsx")
        self.chk_disable_styles.setChecked(True)
        self.chk_turbo_csv.setToolTip("ON: For large merged results, write full data to CSV instead of forcing a very large xlsx. This is usually the fastest and most stable option.")
        self.chk_summary_xlsx.setToolTip("ON: When CSV output is used, also create a small summary xlsx for Excel review. Recommended for large logs.")
        self.chk_fast_xlsx.setToolTip("ON: Use a lighter xlsx writer path and reduce expensive Excel features where possible.")
        self.chk_skip_merged1.setToolTip("ON: Skip the duplicate Merged_1 compatibility sheet. This reduces file size and write time.")
        self.chk_timestamp_text.setToolTip("ON: Write Timestamp as text for speed. OFF uses Excel datetime formatting, which is slower but easier for Excel date filters.")
        self.chk_disable_styles.setToolTip("ON: Disable heavy cell styling for large xlsx files. This improves speed and avoids Excel repair issues.")
        self.worker_combo = QComboBox()
        self.worker_combo.addItem("Auto", "0")
        for n in [2, 4, 6, 8, 12, 16]:
            self.worker_combo.addItem(str(n), str(n))
        self.worker_combo.setToolTip("Parser worker count. Auto is recommended unless you need to limit CPU usage.")
        perf_layout.addWidget(self.chk_turbo_csv)
        perf_layout.addWidget(self.chk_summary_xlsx)
        perf_layout.addWidget(self.chk_fast_xlsx)
        perf_layout.addWidget(self.chk_skip_merged1)
        perf_layout.addWidget(self.chk_timestamp_text)
        perf_layout.addWidget(self.chk_disable_styles)
        perf_layout.addWidget(QLabel("Parser workers"))
        perf_layout.addWidget(self.worker_combo)
        root.addWidget(perf_box)

        noise_group = QGroupBox("Smart Noise Learning / Filter")
        noise_layout = QHBoxLayout(noise_group)
        self.noise_enable_merge_chk = QCheckBox("Apply approved Noise Rules in Merge output")
        self.noise_enable_merge_chk.setChecked(True)
        self.noise_exclude_output_chk = QCheckBox("Legacy: exclude approved noise from output")
        self.noise_exclude_output_chk.setVisible(False)
        self.noise_learning_chk = QCheckBox("Learning Mode: suggest only")
        self.noise_learning_chk.setChecked(True)
        self.noise_manage_btn = QPushButton("Manage Noise Rules")
        self.noise_manage_btn.clicked.connect(self.manage_noise_rules_main)
        self.noise_manage_btn.setToolTip("Open the Noise Rules Manager to enable, disable, test, edit, or delete operator-approved rules.")
        self.noise_enable_merge_chk.setToolTip("ON: Approved and enabled Noise rules are applied during Merge. Matching rows are removed from output. Turn OFF to keep all records even if rules exist.")
        self.noise_learning_chk.setToolTip("ON: The tool may suggest noise candidates, but nothing is removed until an operator approves a rule. Recommended ON.")
        noise_layout.addWidget(self.noise_enable_merge_chk)
        noise_layout.addWidget(self.noise_exclude_output_chk)
        noise_layout.addWidget(self.noise_learning_chk)
        noise_layout.addWidget(self.noise_manage_btn)
        noise_layout.addStretch(1)
        root.addWidget(noise_group)

        btn_layout = QHBoxLayout()
        self.run_btn = QPushButton("Run Log Merge")
        self.run_btn.clicked.connect(self.run_clicked)
        self.pause_btn = QPushButton("Pause")
        self.pause_btn.clicked.connect(self.pause_clicked)
        self.pause_btn.setEnabled(False)
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.cancel_clicked)
        self.cancel_btn.setEnabled(False)
        self.import_psc_btn = QPushButton("Import PSC Only")
        self.import_psc_btn.clicked.connect(self.import_psc_clicked)
        self.import_review_btn = QPushButton("Import Review Only")
        self.import_review_btn.clicked.connect(self.import_review_clicked)
        self.import_file_btn = QPushButton("Import Selected File Only")
        self.import_file_btn.clicked.connect(self.import_selected_file_clicked)
        self.export_project_btn = QPushButton("Export Project")
        self.export_project_btn.clicked.connect(self.export_project_clicked)
        self.import_project_btn = QPushButton("Import Project")
        self.import_project_btn.clicked.connect(self.import_project_clicked)
        self.split_merge_btn = QPushButton("Split Merge Project")
        self.split_merge_btn.clicked.connect(self.split_merge_clicked)
        self.split_merge_btn.setToolTip("Split a merged Search xlsx/csv back into per-source reconstructed log files. Uses Raw column grouped by SourceType and File.")
        self.open_out_btn = QPushButton("Open Output Folder")
        self.open_out_btn.clicked.connect(self.open_output_folder)
        self.update_file_type_btn = QPushButton("Update File Type")
        self.update_file_type_btn.clicked.connect(self.update_file_type_clicked)
        self.update_file_type_btn.setToolTip("Install or update a File Type plugin ZIP. New log types can be added without rebuilding the EXE.")
        self.reset_defaults_btn = QPushButton("Reset to Defaults")
        self.reset_defaults_btn.clicked.connect(self.reset_to_defaults_clicked)
        self.reset_defaults_btn.setToolTip("Reset folders, date range, site information, and log types to blank. Performance is set to the recommended fastest settings. Smart Noise checkboxes are turned ON.")
        self.viewer_btn = QPushButton("Open Dual Log Viewer / Log Explorer")
        self.viewer_btn.clicked.connect(self.open_dual_viewer)
        self.viewer_window = None
        btn_layout.addWidget(self.run_btn)
        btn_layout.addWidget(self.pause_btn)
        btn_layout.addWidget(self.cancel_btn)
        btn_layout.addWidget(self.import_psc_btn)
        btn_layout.addWidget(self.import_review_btn)
        btn_layout.addWidget(self.import_file_btn)
        btn_layout.addWidget(self.export_project_btn)
        btn_layout.addWidget(self.import_project_btn)
        btn_layout.addWidget(self.split_merge_btn)
        btn_layout.addWidget(self.open_out_btn)
        btn_layout.addWidget(self.update_file_type_btn)
        btn_layout.addWidget(self.reset_defaults_btn)
        btn_layout.addWidget(self.viewer_btn)
        root.addLayout(btn_layout)

        self.progress = QProgressBar(); self.progress.setRange(0, 100)
        root.addWidget(self.progress)
        self.status_label = QLabel("Ready")
        root.addWidget(self.status_label)
        self.log_view = QTextEdit(); self.log_view.setReadOnly(True)
        root.addWidget(self.log_view, stretch=1)

    def refresh_plugin_checkboxes(self):
        # Remove old dynamic plugin checkboxes from the layout.
        for chk in list(getattr(self, "plugin_checks", {}).values()):
            try:
                self.type_layout.removeWidget(chk)
                chk.deleteLater()
            except Exception:
                pass
        self.plugin_checks = {}
        for plg in load_file_type_plugins():
            pid = str(plg.get("id", "")).strip()
            if not pid:
                continue
            label = f"{plg.get('display_name', pid)} (Plugin)"
            chk = QCheckBox(label)
            chk.setChecked(False)
            chk.setToolTip(f"Plugin: {pid} / version {plg.get('version','')} / patterns: {', '.join(plg.get('patterns', []))}")
            self.plugin_checks[pid] = chk
            self.type_layout.addWidget(chk)

    def update_file_type_clicked(self):
        dlg = PluginManagerDialog(self)
        dlg.exec()
        self.refresh_plugin_checkboxes()
        self.log("Plugin list reloaded. Newly installed File Types are available as plugin checkboxes and in Smart File Discovery.")

    def log(self, msg: str):
        self.log_view.append(msg)

    def _show_progress_popup(self, title: str, label: str, minimum: int = 0, maximum: int = 100, cancellable: bool = True):
        self._close_progress_popup()
        dlg = QProgressDialog(label, "Cancel" if cancellable else None, minimum, maximum, self)
        dlg.setWindowTitle(title)
        dlg.setMinimumDuration(700)
        dlg.setAutoClose(False)
        dlg.setAutoReset(False)
        dlg.setWindowModality(Qt.WindowModal)
        if cancellable:
            dlg.canceled.connect(self.cancel_clicked)
        dlg.show()
        QApplication.processEvents()
        self.progress_dialog = dlg
        return dlg

    def _close_progress_popup(self):
        try:
            if self.progress_dialog is not None:
                self.progress_dialog.close()
        except Exception:
            pass
        self.progress_dialog = None

    def _run_import_task(self, title: str, label: str, func, done_message_prefix: str):
        self._show_progress_popup(title, label, 0, 0, cancellable=False)
        self.status_label.setText(label)
        self.import_worker = SimpleTaskWorker(func)
        def ok(out):
            self._close_progress_popup()
            self.import_worker = None
            self.status_label.setText(f"{done_message_prefix} completed")
            self.log(f"{done_message_prefix} completed: {out}")
            QMessageBox.information(self, APP_TITLE, f"{done_message_prefix} completed.\n\nOutput:\n{out}")
        def failed(text):
            self._close_progress_popup()
            self.import_worker = None
            self.status_label.setText(f"{done_message_prefix} failed")
            self.log(text)
            QMessageBox.critical(self, APP_TITLE, text)
        self.import_worker.finished_ok.connect(ok)
        self.import_worker.failed.connect(failed)
        self.import_worker.start()

    def pick_source(self):
        d = QFileDialog.getExistingDirectory(self, "Select Source Log Folder", self.source_edit.text() or str(Path.home()))
        if d:
            self.source_edit.setText(d)
            if not self.output_edit.text():
                self.output_edit.setText(d)

    def pick_output(self):
        d = QFileDialog.getExistingDirectory(self, "Select Output Folder", self.output_edit.text() or self.source_edit.text() or str(Path.home()))
        if d:
            self.output_edit.setText(d)

    def sync_site_from_serial(self, text: str):
        for item in self.site_map:
            if str(item.get("serial", "")) == text and self.site_combo.currentText() != str(item.get("site", "")):
                self.site_combo.setCurrentText(str(item.get("site", "")))
                break

    def sync_serial_from_site(self, text: str):
        for item in self.site_map:
            if str(item.get("site", "")) == text and self.serial_combo.currentText() != str(item.get("serial", "")):
                self.serial_combo.setCurrentText(str(item.get("serial", "")))
                break

    def collect_options(self) -> RunOptions:
        return RunOptions(
            source_folder=self.source_edit.text().strip(),
            output_folder=self.output_edit.text().strip() or self.source_edit.text().strip(),
            recursive=self.recursive_chk.isChecked(),
            use_start=self.use_start.isChecked(),
            use_end=self.use_end.isChecked(),
            start_date=self.start_date.date().toString("yyyy-MM-dd"),
            end_date=self.end_date.date().toString("yyyy-MM-dd"),
            serial=self.serial_combo.currentText().strip(),
            site=self.site_combo.currentText().strip(),
            include_ws=self.chk_ws.isChecked(),
            include_cga=self.chk_cga.isChecked(),
            include_csa=self.chk_csa.isChecked(),
            include_mrserver=self.chk_mrserver.isChecked(),
            include_gesys=self.chk_gesys.isChecked(),
            include_lais=self.chk_lais.isChecked(),
            include_psc=self.chk_psc.isChecked(),
            include_unknown=self.chk_unknown.isChecked(),
            turbo_csv=self.chk_turbo_csv.isChecked(),
            summary_xlsx_only=self.chk_summary_xlsx.isChecked(),
            fast_xlsx_mode=self.chk_fast_xlsx.isChecked(),
            csv_summary_mode=self.chk_summary_xlsx.isChecked(),
            skip_merged1=self.chk_skip_merged1.isChecked(),
            timestamp_as_text=self.chk_timestamp_text.isChecked(),
            disable_styles_large=self.chk_disable_styles.isChecked(),
            worker_count=int(self.worker_combo.currentData() or 0),
            noise_enabled_merge=self.noise_enable_merge_chk.isChecked(),
            noise_exclude_output=(self.noise_exclude_output_chk.isChecked() or self.noise_enable_merge_chk.isChecked()),
            noise_learning=self.noise_learning_chk.isChecked(),
            plugin_types=[pid for pid, chk in getattr(self, "plugin_checks", {}).items() if chk.isChecked()],
        )

    def validate_options(self, opt: RunOptions) -> Optional[str]:
        if not opt.source_folder:
            return "Source Log Folder is required."
        if not Path(opt.source_folder).exists():
            return "Source Log Folder does not exist."
        if opt.use_start and opt.use_end and opt.start_date and opt.end_date and opt.start_date > opt.end_date:
            return "Start Date is later than End Date."
        if not any([opt.include_ws, opt.include_cga, opt.include_csa, opt.include_mrserver, opt.include_gesys, opt.include_lais, opt.include_psc, opt.include_unknown]):
            return "Select at least one log type."
        return None

    def pause_clicked(self):
        if not self.worker:
            return
        paused = self.pause_btn.text() == "Pause"
        self.worker.set_paused(paused)
        if paused:
            self.pause_btn.setText("Resume")
            self.status_label.setText("Paused")
            self.log("Paused by user.")
        else:
            self.pause_btn.setText("Pause")
            self.status_label.setText("Running...")
            self.log("Resumed by user.")

    def cancel_clicked(self):
        if not self.worker:
            return
        self.worker.request_cancel()
        self.cancel_btn.setEnabled(False)
        self.pause_btn.setEnabled(False)
        self.status_label.setText("Cancelling...")
        self.log("Cancel requested. Current file/write chunk will finish, then the operation will stop.")

    def import_psc_clicked(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select PSC log", self.source_edit.text() or str(Path.home()), "PSC Log (*.log *.txt);;All Files (*.*)")
        if not path:
            return
        out_dir = self.output_edit.text().strip() or str(Path(path).parent)
        self._run_import_task(
            "PSC Import Progress",
            "Importing PSC and writing output...",
            lambda: import_psc_only(path, out_dir),
            "PSC import",
        )

    def import_review_clicked(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Review file", self.source_edit.text() or str(Path.home()), "Review files (review.out review.out.ar review.out.*);;All Files (*.*)")
        if not path:
            return
        if not is_review_file(Path(path).name):
            QMessageBox.warning(self, APP_TITLE, "Review Import target must be review.out, review.out.ar, or review.out.*")
            return
        out_dir = self.output_edit.text().strip() or str(Path(path).parent)
        self._run_import_task(
            "Review Import Progress",
            "Importing review.out and writing output...",
            lambda: import_reviewout_only(path, out_dir),
            "Review import",
        )

    def import_selected_file_clicked(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select one log file to import only",
            self.source_edit.text() or str(Path.home()),
            "Log files (*.log *.txt *.out *.*);;All Files (*.*)",
        )
        if not path:
            return
        out_dir = self.output_edit.text().strip() or str(Path(path).parent)
        self._run_import_task(
            "Selected File Import Progress",
            "Importing selected file and writing output...",
            lambda: import_selected_file_only(path, out_dir),
            "Selected file import",
        )

    def current_project_dict(self) -> dict[str, Any]:
        return {
            "app_version": APP_VERSION,
            "source_folder": self.source_edit.text(),
            "output_folder": self.output_edit.text(),
            "recursive": self.recursive_chk.isChecked(),
            "use_start": self.use_start.isChecked(),
            "use_end": self.use_end.isChecked(),
            "start_date": self.start_date.date().toString("yyyy-MM-dd"),
            "end_date": self.end_date.date().toString("yyyy-MM-dd"),
            "serial": self.serial_combo.currentText(),
            "site": self.site_combo.currentText(),
            "log_types": {
                "WS": self.chk_ws.isChecked(),
                "CGA": self.chk_cga.isChecked(),
                "CSA": self.chk_csa.isChecked(),
                "MRSERVER": self.chk_mrserver.isChecked(),
                "GESYS": self.chk_gesys.isChecked(),
                "LAIS": self.chk_lais.isChecked(),
                "PSC": self.chk_psc.isChecked(),
                "UNKNOWN": self.chk_unknown.isChecked(),
            },
            "performance": {
                "turbo_csv": self.chk_turbo_csv.isChecked(),
                "summary_xlsx_only": self.chk_summary_xlsx.isChecked(),
                "worker_count": self.worker_combo.currentData(),
            },
            "noise": {
                "enable_merge": self.noise_enable_merge_chk.isChecked(),
                "exclude_output": self.noise_exclude_output_chk.isChecked(),
                "learning": self.noise_learning_chk.isChecked(),
            },
        }

    def apply_project_dict(self, data: dict[str, Any]):
        self.source_edit.setText(str(data.get("source_folder", "")))
        self.output_edit.setText(str(data.get("output_folder", "")))
        self.recursive_chk.setChecked(bool(data.get("recursive", True)))
        self.use_start.setChecked(bool(data.get("use_start", False)))
        self.use_end.setChecked(bool(data.get("use_end", False)))
        try:
            self.start_date.setDate(QDate.fromString(str(data.get("start_date", "")), "yyyy-MM-dd"))
        except Exception:
            pass
        try:
            self.end_date.setDate(QDate.fromString(str(data.get("end_date", "")), "yyyy-MM-dd"))
        except Exception:
            pass
        self.serial_combo.setCurrentText(str(data.get("serial", "")))
        self.site_combo.setCurrentText(str(data.get("site", "")))
        lt = data.get("log_types", {}) or {}
        self.chk_ws.setChecked(bool(lt.get("WS", True)))
        self.chk_cga.setChecked(bool(lt.get("CGA", True)))
        self.chk_csa.setChecked(bool(lt.get("CSA", True)))
        self.chk_mrserver.setChecked(bool(lt.get("MRSERVER", True)))
        self.chk_gesys.setChecked(bool(lt.get("GESYS", True)))
        self.chk_lais.setChecked(bool(lt.get("LAIS", True)))
        self.chk_psc.setChecked(bool(lt.get("PSC", True)))
        self.chk_unknown.setChecked(bool(lt.get("UNKNOWN", False)))
        perf = data.get("performance", {}) or {}
        self.chk_turbo_csv.setChecked(bool(perf.get("turbo_csv", True)))
        self.chk_summary_xlsx.setChecked(bool(perf.get("summary_xlsx_only", True)))
        self.chk_fast_xlsx.setChecked(bool(perf.get("fast_xlsx_mode", True)))
        self.chk_skip_merged1.setChecked(bool(perf.get("skip_merged1", True)))
        self.chk_timestamp_text.setChecked(bool(perf.get("timestamp_as_text", True)))
        self.chk_disable_styles.setChecked(bool(perf.get("disable_styles_large", True)))
        wc = str(perf.get("worker_count", "0"))
        idx = self.worker_combo.findData(wc)
        if idx >= 0:
            self.worker_combo.setCurrentIndex(idx)
        noise = data.get("noise", {}) or {}
        self.noise_enable_merge_chk.setChecked(bool(noise.get("enable_merge", False)))
        self.noise_exclude_output_chk.setChecked(bool(noise.get("exclude_output", False)))
        self.noise_learning_chk.setChecked(bool(noise.get("learning", True)))

    def export_project_clicked(self):
        default_dir = self.output_edit.text().strip() or self.source_edit.text().strip() or str(Path.home())
        default_path = str(Path(default_dir) / f"LogMergeProject_{datetime.now().strftime('%Y%m%d_%H%M%S')}.lmpkg")
        out_path, _ = QFileDialog.getSaveFileName(self, "Export Project", default_path, "Log Merge Project (*.lmpkg);;Zip files (*.zip);;All Files (*.*)")
        if not out_path:
            return
        data = self.current_project_dict()
        try:
            with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as z:
                z.writestr("project.json", json.dumps(data, ensure_ascii=False, indent=2))
                dbp = noise_db_path()
                if dbp.exists():
                    z.write(dbp, "NoiseDB.sqlite")
            QMessageBox.information(self, APP_TITLE, f"Project exported.\n\n{out_path}")
        except Exception:
            text = traceback.format_exc()
            self.log(text)
            QMessageBox.critical(self, APP_TITLE, text)

    def import_project_clicked(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import Project", str(Path.home()), "Log Merge Project (*.lmpkg *.zip);;All Files (*.*)")
        if not path:
            return
        try:
            with zipfile.ZipFile(path, "r") as z:
                data = json.loads(z.read("project.json").decode("utf-8"))
                self.apply_project_dict(data)
                if "NoiseDB.sqlite" in z.namelist():
                    if QMessageBox.question(self, APP_TITLE, "Import NoiseDB.sqlite from this project and replace current rules?") == QMessageBox.Yes:
                        noise_db_path().write_bytes(z.read("NoiseDB.sqlite"))
            self.save_settings()
            QMessageBox.information(self, APP_TITLE, "Project imported.")
        except Exception:
            text = traceback.format_exc()
            self.log(text)
            QMessageBox.critical(self, APP_TITLE, text)

    def split_merge_clicked(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select merged Search xlsx/csv to split",
            self.output_edit.text().strip() or self.source_edit.text().strip() or str(Path.home()),
            "Merged output (*.xlsx *.csv);;All Files (*.*)",
        )
        if not path:
            return
        default_out = str(Path(path).with_suffix("").with_name(Path(path).stem + "_split"))
        out_dir = QFileDialog.getExistingDirectory(self, "Select split output folder", default_out)
        if not out_dir:
            return
        self._run_import_task(
            "Split Merge Project Progress",
            "Splitting merged output into per-source reconstructed logs...",
            lambda: split_merge_project(path, out_dir, log_cb=self.log),
            "Split merge project",
        )

    def run_clicked(self):
        opt = self.collect_options()
        # v33 Smart File Discovery: log types can be blank before scanning.
        # First validate only the source/date basics, then let the discovery
        # dialog detect log types and let the operator choose final files.
        if not opt.source_folder:
            QMessageBox.warning(self, APP_TITLE, "Source Log Folder is required.")
            return
        if not Path(opt.source_folder).exists():
            QMessageBox.warning(self, APP_TITLE, "Source Log Folder does not exist.")
            return
        if opt.use_start and opt.use_end and opt.start_date and opt.end_date and opt.start_date > opt.end_date:
            QMessageBox.warning(self, APP_TITLE, "Start Date is later than End Date.")
            return

        dlg = SmartFileDiscoveryDialog(self, opt)
        if dlg.exec() != QDialog.Accepted:
            self.status_label.setText("Cancelled before merge")
            return

        selected_types = set(dlg.selected_types)
        opt.include_ws = "WS" in selected_types
        opt.include_cga = "CGA" in selected_types
        opt.include_csa = "CSA" in selected_types
        opt.include_mrserver = "MRSERVER" in selected_types
        opt.include_gesys = "GESYS" in selected_types
        opt.include_lais = "LAIS" in selected_types
        opt.include_psc = "PSC" in selected_types
        opt.include_unknown = "UNKNOWN" in selected_types
        opt.selected_files = dlg.selected_files
        if not enabled_types(opt):
            QMessageBox.warning(self, APP_TITLE, "No merge-capable log type is selected. Review files are Import Only.")
            return

        self.save_settings()
        self.log_view.clear()
        self.log(f"Smart File Discovery selected {len(opt.selected_files or [])} files.")
        self.progress.setValue(0)
        self.status_label.setText("Running...")
        self.run_btn.setEnabled(False)
        self.pause_btn.setEnabled(True)
        self.cancel_btn.setEnabled(True)
        self.pause_btn.setText("Pause")
        self._show_progress_popup("Log Merge Progress", "Parsing selected files...", 0, 100, cancellable=True)
        self.worker = MergeWorker(opt)
        self.worker.progress.connect(self.on_progress)
        self.worker.log.connect(self.log)
        self.worker.finished_ok.connect(self.on_finished)
        self.worker.failed.connect(self.on_failed)
        self.worker.start()

    def on_progress(self, idx: int, total: int, name: str):
        pct = max(0, min(100, int(idx * 100 / total))) if total else 0
        self.progress.setValue(pct)
        self.progress.setFormat(f"{pct}%")
        self.status_label.setText(f"{pct}%  ({idx}/{total}): {name}")
        if self.progress_dialog is not None:
            self.progress_dialog.setRange(0, 100)
            self.progress_dialog.setValue(pct)
            self.progress_dialog.setLabelText(f"{pct}%  ({idx}/{total})\n{name}")
            QApplication.processEvents()

    def on_finished(self, result: RunResult):
        self._close_progress_popup()
        self.run_btn.setEnabled(True)
        self.pause_btn.setEnabled(False)
        self.cancel_btn.setEnabled(False)
        self.progress.setValue(100)
        self.status_label.setText("Completed")
        self.log("Completed.")
        self.log(f"Output: {result.output_path}")
        if result.csa_error_count > 0:
            self.log(f"CSA errors detected: {result.csa_error_count}")
        QMessageBox.information(self, APP_TITLE, f"Completed.\n\nOutput:\n{result.output_path}")

    def on_failed(self, text: str):
        self._close_progress_popup()
        self.run_btn.setEnabled(True)
        self.pause_btn.setEnabled(False)
        self.cancel_btn.setEnabled(False)
        self.status_label.setText("Failed")
        self.log(text)
        QMessageBox.critical(self, APP_TITLE, text)

    def open_output_folder(self):
        p = self.output_edit.text().strip() or self.source_edit.text().strip()
        if p and Path(p).exists():
            os.startfile(p)  # type: ignore[attr-defined]

    def open_dual_viewer(self):
        if self.viewer_window is None:
            self.viewer_window = DualLogViewer(self)
        self.viewer_window.show()
        self.viewer_window.raise_()
        self.viewer_window.activateWindow()

    def apply_default_settings(self):
        """Apply factory defaults without clearing saved settings directly."""
        self.source_edit.setText("")
        self.output_edit.setText("")
        self.recursive_chk.setChecked(True)
        self.use_start.setChecked(False)
        self.use_end.setChecked(False)
        self.start_date.setDate(QDate.currentDate())
        self.end_date.setDate(QDate.currentDate())
        self.serial_combo.setCurrentText("")
        self.site_combo.setCurrentText("")
        for chk in [self.chk_ws, self.chk_cga, self.chk_csa, self.chk_mrserver, self.chk_gesys, self.chk_lais, self.chk_psc, self.chk_review, self.chk_unknown]:
            chk.setChecked(False)
        for chk in getattr(self, "plugin_checks", {}).values():
            chk.setChecked(False)
        # Recommended fastest/stablest defaults.
        self.chk_turbo_csv.setChecked(True)
        self.chk_summary_xlsx.setChecked(True)
        self.chk_fast_xlsx.setChecked(True)
        self.chk_skip_merged1.setChecked(True)
        self.chk_timestamp_text.setChecked(True)
        self.chk_disable_styles.setChecked(True)
        idx = self.worker_combo.findData("0")
        if idx >= 0:
            self.worker_combo.setCurrentIndex(idx)
        # Smart Noise defaults: both visible options ON.
        self.noise_enable_merge_chk.setChecked(True)
        self.noise_exclude_output_chk.setChecked(False)
        self.noise_learning_chk.setChecked(True)

    def reset_to_defaults_clicked(self):
        reply = QMessageBox.question(
            self,
            APP_TITLE,
            "Reset UI settings to defaults?\n\nFolders, date range, site information, and log type selections will be cleared.\nPerformance will use the recommended fastest settings.\nSmart Noise options will be turned ON.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        self.apply_default_settings()
        self.save_settings()
        self.status_label.setText("Defaults restored")
        self.log("Settings were reset to defaults.")

    def manage_noise_rules_main(self):
        dlg = NoiseRuleManagerDialog(self)
        dlg.exec()

    def load_settings(self):
        # Restore the previous user state. If no previous state exists, use factory defaults.
        self.source_edit.setText(self.settings.value("source_folder", ""))
        self.output_edit.setText(self.settings.value("output_folder", ""))
        self.recursive_chk.setChecked(self.settings.value("recursive", "true") == "true")
        self.use_start.setChecked(self.settings.value("use_start_date", "false") == "true")
        self.use_end.setChecked(self.settings.value("use_end_date", "false") == "true")
        sd = self.settings.value("start_date", "")
        ed = self.settings.value("end_date", "")
        if sd:
            self.start_date.setDate(QDate.fromString(sd, "yyyy-MM-dd"))
        if ed:
            self.end_date.setDate(QDate.fromString(ed, "yyyy-MM-dd"))
        self.serial_combo.setCurrentText(self.settings.value("serial", ""))
        self.site_combo.setCurrentText(self.settings.value("site", ""))
        self.chk_ws.setChecked(self.settings.value("log_ws", "false") == "true")
        self.chk_cga.setChecked(self.settings.value("log_cga", "false") == "true")
        self.chk_csa.setChecked(self.settings.value("log_csa", "false") == "true")
        self.chk_mrserver.setChecked(self.settings.value("log_mrserver", "false") == "true")
        self.chk_gesys.setChecked(self.settings.value("log_gesys", "false") == "true")
        self.chk_lais.setChecked(self.settings.value("log_lais", "false") == "true")
        self.chk_psc.setChecked(self.settings.value("log_psc", "false") == "true")
        self.chk_review.setChecked(self.settings.value("log_review", "false") == "true")
        self.chk_unknown.setChecked(self.settings.value("log_unknown", "false") == "true")
        saved_plugins = set(str(self.settings.value("log_plugins", "")).split(","))
        for pid, chk in getattr(self, "plugin_checks", {}).items():
            chk.setChecked(pid in saved_plugins)
        self.chk_turbo_csv.setChecked(self.settings.value("turbo_csv", "true") == "true")
        self.chk_summary_xlsx.setChecked(self.settings.value("summary_xlsx", "true") == "true")
        self.chk_fast_xlsx.setChecked(self.settings.value("fast_xlsx", "true") == "true")
        self.chk_skip_merged1.setChecked(self.settings.value("skip_merged1", "true") == "true")
        self.chk_timestamp_text.setChecked(self.settings.value("timestamp_text", "true") == "true")
        self.chk_disable_styles.setChecked(self.settings.value("disable_styles", "true") == "true")
        if hasattr(self, "noise_enable_merge_chk"):
            self.noise_enable_merge_chk.setChecked(self.settings.value("noise_enable_merge", "true") == "true")
            self.noise_exclude_output_chk.setChecked(self.settings.value("noise_exclude_output", "false") == "true")
            self.noise_learning_chk.setChecked(self.settings.value("noise_learning", "true") == "true")
        worker_value = self.settings.value("worker_count", "0")
        idx = self.worker_combo.findData(str(worker_value))
        if idx >= 0:
            self.worker_combo.setCurrentIndex(idx)

    def save_settings(self):
        self.settings.setValue("source_folder", self.source_edit.text())
        self.settings.setValue("output_folder", self.output_edit.text())
        self.settings.setValue("recursive", "true" if self.recursive_chk.isChecked() else "false")
        self.settings.setValue("use_start_date", "true" if self.use_start.isChecked() else "false")
        self.settings.setValue("use_end_date", "true" if self.use_end.isChecked() else "false")
        self.settings.setValue("start_date", self.start_date.date().toString("yyyy-MM-dd"))
        self.settings.setValue("end_date", self.end_date.date().toString("yyyy-MM-dd"))
        self.settings.setValue("noise_enable_merge", "true" if self.noise_enable_merge_chk.isChecked() else "false")
        self.settings.setValue("noise_exclude_output", "true" if self.noise_exclude_output_chk.isChecked() else "false")
        self.settings.setValue("noise_learning", "true" if self.noise_learning_chk.isChecked() else "false")
        self.settings.setValue("serial", self.serial_combo.currentText())
        self.settings.setValue("site", self.site_combo.currentText())
        self.settings.setValue("log_ws", "true" if self.chk_ws.isChecked() else "false")
        self.settings.setValue("log_cga", "true" if self.chk_cga.isChecked() else "false")
        self.settings.setValue("log_csa", "true" if self.chk_csa.isChecked() else "false")
        self.settings.setValue("log_mrserver", "true" if self.chk_mrserver.isChecked() else "false")
        self.settings.setValue("log_gesys", "true" if self.chk_gesys.isChecked() else "false")
        self.settings.setValue("log_lais", "true" if self.chk_lais.isChecked() else "false")
        self.settings.setValue("log_psc", "true" if self.chk_psc.isChecked() else "false")
        self.settings.setValue("log_review", "true" if self.chk_review.isChecked() else "false")
        self.settings.setValue("log_unknown", "true" if self.chk_unknown.isChecked() else "false")
        self.settings.setValue("log_plugins", ",".join([pid for pid, chk in getattr(self, "plugin_checks", {}).items() if chk.isChecked()]))
        self.settings.setValue("turbo_csv", "true" if self.chk_turbo_csv.isChecked() else "false")
        self.settings.setValue("summary_xlsx", "true" if self.chk_summary_xlsx.isChecked() else "false")
        self.settings.setValue("fast_xlsx", "true" if self.chk_fast_xlsx.isChecked() else "false")
        self.settings.setValue("skip_merged1", "true" if self.chk_skip_merged1.isChecked() else "false")
        self.settings.setValue("timestamp_text", "true" if self.chk_timestamp_text.isChecked() else "false")
        self.settings.setValue("disable_styles", "true" if self.chk_disable_styles.isChecked() else "false")
        self.settings.setValue("worker_count", str(self.worker_combo.currentData() or "0"))
        self.settings.sync()


def show_startup_error(text: str) -> None:
    write_startup_log(text)
    try:
        _app = QApplication.instance() or QApplication(sys.argv)
        QMessageBox.critical(None, APP_TITLE, text + f"\n\nLog:\n{startup_log_path()}")
    except Exception:
        pass


def main():
    def excepthook(exc_type, exc, tb):
        text = "Unhandled application error.\n\n" + "".join(traceback.format_exception(exc_type, exc, tb))
        show_startup_error(text)
    sys.excepthook = excepthook

    try:
        app = QApplication(sys.argv)
        w = MainWindow()
        w.show()
        sys.exit(app.exec())
    except Exception:
        show_startup_error("Startup failed.\n\n" + traceback.format_exc())
        raise


if __name__ == "__main__":
    main()
